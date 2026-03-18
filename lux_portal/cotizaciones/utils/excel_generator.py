#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de cotizaciones Excel para FreightWise
Genera documentos en español e ingles
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import re
import io

# Ruta del logo (relativa al proyecto)
def get_logo_path():
    """Obtiene la ruta del logo de forma compatible con cualquier entorno."""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(current_dir, '..', '..', 'static', 'images', 'freightwise_logo.png')
    return os.path.normpath(logo_path)

# Intentar importar librerias de traduccion
try:
    from deep_translator import GoogleTranslator
    from langdetect import detect
    TRADUCCION_DISPONIBLE = True
except ImportError:
    TRADUCCION_DISPONIBLE = False

# Diccionario de dias de la semana abreviados
DIAS_ES_EN = {
    'LUN': 'MON', 'MAR': 'TUE', 'MIE': 'WED', 'MIER': 'WED',
    'JUE': 'THU', 'VIE': 'FRI', 'SAB': 'SAT', 'DOM': 'SUN'
}
DIAS_EN_ES = {v: k for k, v in DIAS_ES_EN.items()}


def traducir_dias(texto, a_ingles=True):
    """Traduce las abreviaciones de dias de la semana en un texto."""
    if not texto or not texto.strip():
        return texto

    resultado = texto
    diccionario = DIAS_ES_EN if a_ingles else DIAS_EN_ES

    for dia_origen, dia_destino in diccionario.items():
        resultado = re.sub(r'\b' + re.escape(dia_origen) + r'\b', dia_destino, resultado, flags=re.IGNORECASE)

    return resultado


def detectar_idioma(texto):
    """Detecta automaticamente el idioma del texto."""
    if not texto or not texto.strip():
        return 'es'

    if not TRADUCCION_DISPONIBLE:
        texto_lower = texto.lower()
        palabras_en = ['the', 'and', 'or', 'is', 'are', 'rate', 'charge']
        palabras_es = ['el', 'la', 'y', 'o', 'es', 'tarifa', 'cargo']

        count_en = sum(1 for palabra in palabras_en if palabra in texto_lower.split())
        count_es = sum(1 for palabra in palabras_es if palabra in texto_lower.split())

        return 'en' if count_en > count_es else 'es'

    try:
        return detect(texto)
    except:
        return 'es'


def traducir_texto_auto(texto, idioma_destino='en'):
    """Traduce texto usando Google Translator."""
    if not texto or not texto.strip() or not TRADUCCION_DISPONIBLE:
        return texto

    try:
        idioma_origen = detectar_idioma(texto)
        if idioma_origen == idioma_destino:
            return texto

        return GoogleTranslator(source=idioma_origen, target=idioma_destino).translate(texto)
    except:
        return texto


def auto_ajustar_columnas_y_filas(ws, fila_inicio, fila_fin):
    """Ajusta automaticamente el ancho de columnas y alto de filas."""
    from openpyxl.cell.cell import MergedCell

    for col_idx in range(1, 16):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for row_idx in range(fila_inicio, fila_fin + 1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value:
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else len(cell_value)
                    if max_line_length > max_length:
                        max_length = max_line_length
            except:
                pass

        if max_length > 0:
            if column_letter in ['L', 'O']:
                adjusted_width = min(max(max_length + 2, 15), 65)
            else:
                adjusted_width = min(max(max_length + 2, 10), 35)
            ws.column_dimensions[column_letter].width = adjusted_width

    for row_idx in range(fila_inicio, fila_fin + 1):
        max_lines = 1
        for col_idx in range(1, 16):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    if lines > max_lines:
                        max_lines = lines
            except:
                pass

        altura = 20 if max_lines <= 1 else 18 + (max_lines - 1) * 12
        ws.row_dimensions[row_idx].height = altura


def generar_hoja(ws, datos, idioma='es'):
    """Genera una hoja de cotizacion en el idioma especificado."""

    # Quitar gridlines
    ws.sheet_view.showGridLines = False

    # Configurar ancho de columnas
    anchos = {'A': 20, 'B': 15, 'C': 30, 'D': 15, 'E': 18, 'F': 18, 'G': 18,
              'H': 12, 'I': 15, 'J': 12, 'K': 12, 'L': 35, 'M': 3, 'N': 12, 'O': 50}
    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    # Fila 1: logo centrado, fila 2: espacio antes del bloque morado
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 6

    # Logo centrado en fila 1 (anclado en G1 ≈ centro horizontal del documento)
    logo_path = get_logo_path()
    if os.path.exists(logo_path):
        img_logo = Image(logo_path)
        img_logo.width = 230
        img_logo.height = 42
        ws.add_image(img_logo, 'G1')

    # Textos segun idioma
    if idioma == 'es':
        titulo = 'C O T I Z A C I O N   F L E T E   A E R E O'
        contacto_label = 'Contacto FreightWise'
        valido_label = 'Valido desde'
        cliente_label = 'CLIENTE'
        mercancia_label = 'MERCANCIA'
        encabezados = ['AEROLINEA', 'VUELO', 'ITINERARIO', 'TIEMPO\nTRANSITO', 'FINCAS\nENTREGA',
                       'SALIDA', 'LLEGADA', 'KG', 'TARIFA', 'AUMENTO\nTARIFA', 'FECHA',
                       'CARGOS ADICIONALES', '', '', 'NOTAS']
    else:
        titulo = 'A I R F R E I G H T   Q U O T A T I O N'
        contacto_label = 'FreightWise Contact'
        valido_label = 'Valid from'
        cliente_label = 'CUSTOMER'
        mercancia_label = 'COMMODITY'
        encabezados = ['AIRLINE', 'FLIGHT', 'ITINERARY', 'TRANSIT\nTIME', 'FARMS\nDELIVER',
                       'DEPARTURE', 'ARRIVAL', 'KG', 'RATE', 'RATE\nINCREASE', 'DATE',
                       'ADD CHARGES', '', '', 'NOTES']

    # Titulo (debajo del logo, fila 1 queda libre para que el logo sea el header visual)
    ws['A1'] = titulo
    ws['A1'].font = Font(name='Arial', size=11, bold=True, color='5F4DCE')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:E1')

    # Color morado único para toda la cabecera (igual que PDF)
    PURPLE = '4535A8'
    purple_fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type='solid')
    white_text = Font(name='Arial', size=9, color='FFFFFF')
    white_bold = Font(name='Arial', size=9, bold=True, color='FFFFFF')

    # Bloque info cliente + contacto — fondo morado, todo centrado (filas 3-6)
    for fila in range(3, 7):
        for col in range(1, 16):
            ws.cell(row=fila, column=col).fill = purple_fill
        ws.row_dimensions[fila].height = 20

    center_style = Alignment(horizontal='center', vertical='center')

    # Lado izquierdo: A-I merged por fila, label+valor centrado
    ws.merge_cells('A3:I3')
    ws['A3'] = f'{cliente_label}:   {datos.get("customer", "")}'
    ws['A3'].font = white_bold
    ws['A3'].alignment = center_style

    ws.merge_cells('A4:I4')
    ws['A4'] = f'ATTN:   {datos.get("attn", "")}'
    ws['A4'].font = white_text
    ws['A4'].alignment = center_style

    ws.merge_cells('A5:I5')
    ws['A5'] = f'{mercancia_label}:   {datos.get("mercancia", "")}'
    ws['A5'].font = white_bold
    ws['A5'].alignment = center_style

    ws.merge_cells('A6:I6')
    ws['A6'] = f'ORI - DES:   {datos.get("ruta", "")}'
    ws['A6'].font = white_bold
    ws['A6'].alignment = center_style

    # Lado derecho: L-O merged por fila, label+valor centrado
    ws.merge_cells('L3:O3')
    ws['L3'] = f'{contacto_label}:   {datos.get("contacto_nombre", "")}'
    ws['L3'].font = white_bold
    ws['L3'].alignment = center_style

    ws.merge_cells('L4:O4')
    ws['L4'] = f'eMail:   {datos.get("contacto_email", "")}'
    ws['L4'].font = white_text
    ws['L4'].alignment = center_style

    ws.merge_cells('L5:O5')
    ws['L5'] = f'{valido_label}:   {datos.get("valid_from", datetime.now().strftime("%m/%d/%Y"))}'
    ws['L5'].font = white_text
    ws['L5'].alignment = center_style

    # Borde exterior del bloque
    thin_info = Side(style='thin', color='7B6ED6')
    ws.cell(row=3, column=1).border  = Border(top=thin_info, left=thin_info)
    ws.cell(row=3, column=15).border = Border(top=thin_info, right=thin_info)
    ws.cell(row=6, column=1).border  = Border(bottom=thin_info, left=thin_info)
    ws.cell(row=6, column=15).border = Border(bottom=thin_info, right=thin_info)
    for col in range(2, 15):
        ws.cell(row=3, column=col).border = Border(top=thin_info)
        ws.cell(row=6, column=col).border = Border(bottom=thin_info)
    for fila in range(4, 6):
        ws.cell(row=fila, column=1).border  = Border(left=thin_info)
        ws.cell(row=fila, column=15).border = Border(right=thin_info)

    # Ruta
    fila_ruta = 8
    ws[f'A{fila_ruta}'] = datos.get('ruta', '')
    ws[f'A{fila_ruta}'].font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    ws[f'A{fila_ruta}'].fill = purple_fill
    ws.merge_cells(f'A{fila_ruta}:O{fila_ruta}')
    ws[f'A{fila_ruta}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[fila_ruta].height = 25

    # Encabezados de tabla
    fila_enc = 9
    header_fill = purple_fill
    white_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    for idx, enc in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_enc, column=idx)
        celda.value = enc
        celda.font = white_font
        celda.fill = header_fill
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border = border

    ws.merge_cells(f'L{fila_enc}:N{fila_enc}')
    ws.row_dimensions[fila_enc].height = 30

    # Datos de la cotizacion
    fila_datos = 10
    aerolineas = datos.get('aerolineas', [])

    # Ocultar columnas J y K si ninguna aerolinea tiene rate increases
    show_ri_cols = any(aero.get('rate_increases') for aero in aerolineas)
    if not show_ri_cols:
        ws.column_dimensions['J'].hidden = True
        ws.column_dimensions['K'].hidden = True

    # Pre-calcular filas por aerolinea
    filas_por_aerolinea = []
    for aero_datos in aerolineas:
        cargos = aero_datos.get('cargos_adicionales', [])
        num_filas = max(4, len(cargos))
        filas_por_aerolinea.append(num_filas)

    # Colores alternos - gris claro / blanco (igual que PDF)
    color_alterno = ['EBEBEB', 'FFFFFF']
    color_actual_idx = 0

    # Primer paso: fusionar columna A
    idx_aerolinea = 0
    while idx_aerolinea < len(aerolineas):
        aerolinea_datos = aerolineas[idx_aerolinea]
        fila_inicio = fila_datos + sum(filas_por_aerolinea[:idx_aerolinea])

        es_continuacion = aerolinea_datos.get('es_continuacion', False)
        if not es_continuacion and idx_aerolinea > 0:
            color_actual_idx += 1

        color_fondo = color_alterno[color_actual_idx % 2]
        fill_color = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type='solid')

        num_rutas = 1
        if not es_continuacion:
            for i in range(idx_aerolinea + 1, len(aerolineas)):
                if aerolineas[i].get('es_continuacion', False):
                    num_rutas += 1
                else:
                    break

        if not es_continuacion:
            filas_totales = sum(filas_por_aerolinea[idx_aerolinea:idx_aerolinea+num_rutas])
            fila_fin = fila_inicio + filas_totales - 1
            ws[f'A{fila_inicio}'] = aerolinea_datos.get('aerolinea', '')
            ws[f'A{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.merge_cells(f'A{fila_inicio}:A{fila_fin}')
            ws[f'A{fila_inicio}'].fill = fill_color
            texto_color = '000000'  # ambos colores Aire/Mar son claros → texto siempre negro
            ws[f'A{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        idx_aerolinea += 1

    # Segundo paso: otras columnas
    for idx_aerolinea, aerolinea_datos in enumerate(aerolineas):
        fila_inicio = fila_datos + sum(filas_por_aerolinea[:idx_aerolinea])

        es_continuacion = aerolinea_datos.get('es_continuacion', False)
        if idx_aerolinea == 0:
            color_idx_temp = 0
        else:
            color_idx_temp = 0
            for i in range(idx_aerolinea):
                if not aerolineas[i].get('es_continuacion', False) and i > 0:
                    color_idx_temp += 1
            if not es_continuacion:
                color_idx_temp += 1

        color_fondo = color_alterno[color_idx_temp % 2]
        fill_color = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type='solid')
        texto_color = '000000'  # Aire/Mar son colores claros → texto siempre negro

        cargos_temp = aerolinea_datos.get('cargos_adicionales', [])
        num_filas_cargos = max(4, len(cargos_temp))

        # Columnas B-G
        columnas_data = [
            ('B', 'vuelo'),
            ('C', 'itinerario'),
            ('D', 'tiempo_transito'),
        ]

        for col, key in columnas_data:
            ws[f'{col}{fila_inicio}'] = aerolinea_datos.get(key, '')
            ws[f'{col}{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'{col}{fila_inicio}:{col}{fila_inicio+num_filas_cargos-1}')
            ws[f'{col}{fila_inicio}'].fill = fill_color
            ws[f'{col}{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # Fincas entrega
        fincas_texto = aerolinea_datos.get('granjas_entrega', '')
        if idioma == 'en':
            fincas_texto = traducir_dias(fincas_texto, a_ingles=True)
        ws[f'E{fila_inicio}'] = fincas_texto
        ws[f'E{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'E{fila_inicio}:E{fila_inicio+num_filas_cargos-1}')
        ws[f'E{fila_inicio}'].fill = fill_color
        ws[f'E{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # Salida
        salida_texto = aerolinea_datos.get('salida', '')
        if idioma == 'en':
            salida_texto = traducir_dias(salida_texto, a_ingles=True)
        ws[f'F{fila_inicio}'] = salida_texto
        ws[f'F{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'F{fila_inicio}:F{fila_inicio+num_filas_cargos-1}')
        ws[f'F{fila_inicio}'].fill = fill_color
        ws[f'F{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # Llegada
        llegada_texto = aerolinea_datos.get('llegada', '')
        if idioma == 'en':
            llegada_texto = traducir_dias(llegada_texto, a_ingles=True)
        ws[f'G{fila_inicio}'] = llegada_texto
        ws[f'G{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'G{fila_inicio}:G{fila_inicio+num_filas_cargos-1}')
        ws[f'G{fila_inicio}'].fill = fill_color
        ws[f'G{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # KG y TARIFA
        kg_rates = aerolinea_datos.get('kg_rates', [])
        if kg_rates:
            kg_text = '\n'.join([kr.get('kg', '') for kr in kg_rates])
            tarifa_text = '\n'.join([f"$ {kr.get('tarifa_cliente', kr.get('tarifa', ''))}" for kr in kg_rates])
        else:
            kg_text = aerolinea_datos.get('kg', '')
            tarifa_text = f"$ {aerolinea_datos.get('tarifa_cliente', aerolinea_datos.get('tarifa', ''))}"

        ws[f'H{fila_inicio}'] = kg_text
        ws[f'H{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'H{fila_inicio}:H{fila_inicio+num_filas_cargos-1}')
        ws[f'H{fila_inicio}'].fill = fill_color
        ws[f'H{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        ws[f'I{fila_inicio}'] = tarifa_text
        ws[f'I{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'I{fila_inicio}:I{fila_inicio+num_filas_cargos-1}')
        ws[f'I{fila_inicio}'].fill = fill_color
        ws[f'I{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # Rate Increases (J = Aumento Tarifa, K = Fecha)
        rate_increases = aerolinea_datos.get('rate_increases', [])
        if rate_increases:
            ri_amounts = '\n'.join([f"${ri.get('amount', '')}" for ri in rate_increases if ri.get('amount')])
            ri_dates_raw = '\n'.join([ri.get('date', '') for ri in rate_increases if ri.get('date')])
        else:
            ri_amounts = ''
            ri_dates_raw = ''

        # Traducir fechas de rate increase
        if ri_dates_raw:
            idioma_ri = detectar_idioma(ri_dates_raw)
            if idioma == 'en' and idioma_ri == 'es':
                ri_dates = traducir_texto_auto(ri_dates_raw, 'en')
                ri_dates = re.sub(r'(\d+)(st|nd|rd|th)\b', r'\1', ri_dates)
            elif idioma == 'es' and idioma_ri == 'en':
                ri_dates = traducir_texto_auto(ri_dates_raw, 'es')
            else:
                ri_dates = ri_dates_raw
        else:
            ri_dates = ri_dates_raw

        ws[f'J{fila_inicio}'] = ri_amounts
        ws[f'J{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'J{fila_inicio}:J{fila_inicio+num_filas_cargos-1}')
        ws[f'J{fila_inicio}'].fill = fill_color
        ws[f'J{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        ws[f'K{fila_inicio}'] = ri_dates
        ws[f'K{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'K{fila_inicio}:K{fila_inicio+num_filas_cargos-1}')
        ws[f'K{fila_inicio}'].fill = fill_color
        ws[f'K{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # Cargos adicionales - una celda combinada por aerolinea (sin líneas extra)
        cargos = aerolinea_datos.get('cargos_adicionales', [])
        if idioma == 'en':
            cargos_traducidos = []
            for cargo in cargos:
                concepto = cargo.get('concepto', '')
                if concepto.lower() == 'fitosanitario':
                    concepto = 'Phytosanitary'
                elif concepto.lower() == 'certificado':
                    concepto = 'Certificate'
                cargos_traducidos.append({'concepto': concepto, 'monto': cargo.get('monto', '')})
            cargos = cargos_traducidos

        # Texto combinado de todos los cargos en una sola celda
        conceptos_combined = '\n'.join([c.get('concepto', '') for c in cargos if c.get('concepto')])
        montos_combined = '\n'.join([f"$ {c.get('monto', '')}" for c in cargos if c.get('monto')])

        fila_fin_cargo = fila_inicio + num_filas_cargos - 1

        # Aplicar fondo a todo el bloque L-N de esta aerolinea
        for row_i in range(fila_inicio, fila_fin_cargo + 1):
            ws[f'L{row_i}'].fill = fill_color
            ws[f'M{row_i}'].fill = fill_color
            ws[f'N{row_i}'].fill = fill_color

        # Combinar L y M-N en una sola celda por aerolinea
        if num_filas_cargos > 1:
            ws.merge_cells(f'L{fila_inicio}:L{fila_fin_cargo}')
            ws.merge_cells(f'M{fila_inicio}:N{fila_fin_cargo}')
        else:
            ws.merge_cells(f'M{fila_inicio}:N{fila_inicio}')

        ws[f'L{fila_inicio}'] = conceptos_combined
        ws[f'L{fila_inicio}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'L{fila_inicio}'].fill = fill_color
        ws[f'L{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        ws[f'M{fila_inicio}'] = montos_combined
        ws[f'M{fila_inicio}'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws[f'M{fila_inicio}'].fill = fill_color
        ws[f'M{fila_inicio}'].font = Font(name='Arial', size=9, color=texto_color)

        # Notas (shifted to O)
        notas_originales = aerolinea_datos.get('notas', '')
        if idioma == 'en':
            idioma_notas = detectar_idioma(notas_originales)
            if idioma_notas == 'es':
                notas = traducir_texto_auto(notas_originales, 'en')
            else:
                notas = notas_originales
        else:
            idioma_notas = detectar_idioma(notas_originales)
            if idioma_notas == 'en':
                notas = traducir_texto_auto(notas_originales, 'es')
            else:
                notas = notas_originales

        ws[f'O{fila_inicio}'] = notas
        ws[f'O{fila_inicio}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'O{fila_inicio}:O{fila_inicio+num_filas_cargos-1}')
        ws[f'O{fila_inicio}'].font = Font(name='Arial', size=8, color=texto_color)
        ws[f'O{fila_inicio}'].fill = fill_color

        # Ajustar altura de filas
        for i in range(fila_inicio, fila_inicio + num_filas_cargos):
            ws.row_dimensions[i].height = 20

    # Tercer paso: fusionar columna O (notas) para grupos de misma aerolinea con nota igual
    idx_aerolinea = 0
    color_actual_idx_m = 0
    while idx_aerolinea < len(aerolineas):
        es_continuacion = aerolineas[idx_aerolinea].get('es_continuacion', False)
        if not es_continuacion and idx_aerolinea > 0:
            color_actual_idx_m += 1

        if not es_continuacion:
            num_rutas = 1
            for i in range(idx_aerolinea + 1, len(aerolineas)):
                if aerolineas[i].get('es_continuacion', False):
                    num_rutas += 1
                else:
                    break

            if num_rutas > 1:
                notas_grupo = [aerolineas[idx_aerolinea + k].get('notas', '') for k in range(num_rutas)]
                if len(set(notas_grupo)) == 1:
                    fila_grupo_inicio = fila_datos + sum(filas_por_aerolinea[:idx_aerolinea])
                    filas_totales = sum(filas_por_aerolinea[idx_aerolinea:idx_aerolinea + num_rutas])
                    fila_grupo_fin = fila_grupo_inicio + filas_totales - 1

                    # Deshacer merges individuales de O en este grupo
                    for k in range(num_rutas):
                        idx_k = idx_aerolinea + k
                        f_ini = fila_datos + sum(filas_por_aerolinea[:idx_k])
                        f_fin = f_ini + filas_por_aerolinea[idx_k] - 1
                        try:
                            ws.unmerge_cells(f'O{f_ini}:O{f_fin}')
                        except (KeyError, ValueError):
                            pass

                    # Merge del grupo completo
                    ws.merge_cells(f'O{fila_grupo_inicio}:O{fila_grupo_fin}')

            idx_aerolinea += num_rutas
        else:
            idx_aerolinea += 1

    # Calcular ultima fila
    ultima_fila = fila_datos + sum(filas_por_aerolinea) - 1 if filas_por_aerolinea else fila_datos

    # Auto-ajustar
    auto_ajustar_columnas_y_filas(ws, fila_datos, ultima_fila)

    # Aplicar bordes a toda la tabla
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for col in range(1, 16):
        ws.cell(row=fila_enc, column=col).border = thin_border
    for row in range(fila_datos, ultima_fila + 1):
        for col in range(1, 16):
            ws.cell(row=row, column=col).border = thin_border

    # Agregar cargos fijos de FreightWise al final
    fila_fw = ultima_fila + 2

    # Titulo de seccion FreightWise - PÚRPURA FREIGHT
    titulo_fw = 'Cargos Adicionales FreightWise:' if idioma == 'es' else 'FreightWise Additional Charges:'
    ws[f'A{fila_fw}'] = titulo_fw
    ws[f'A{fila_fw}'].font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    ws[f'A{fila_fw}'].fill = purple_fill
    ws[f'A{fila_fw}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A{fila_fw}:O{fila_fw}')
    ws.row_dimensions[fila_fw].height = 25

    # Cargos FreightWise desde datos (editables) o defaults
    traducciones_en = {
        'certificado': 'Certificate',
        'fitosanitario': 'Phytosanitary',
    }
    cargos_fw = datos.get('cargos_freightwise', [
        {'concepto': 'Due Agent', 'monto': '50.00'},
        {'concepto': 'Certificado', 'monto': '15.00'},
        {'concepto': 'Fitosanitario', 'monto': '2.50'}
    ])

    fila_cargo_fw = fila_fw + 1
    for cargo in cargos_fw:
        concepto = cargo.get('concepto', '')
        if idioma == 'en':
            concepto = traducciones_en.get(concepto.lower(), concepto)

        # Concepto en columnas E-K (centrado)
        ws[f'E{fila_cargo_fw}'] = concepto
        ws[f'E{fila_cargo_fw}'].font = Font(name='Arial', size=9, color='000000')
        ws[f'E{fila_cargo_fw}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'E{fila_cargo_fw}:K{fila_cargo_fw}')

        # Monto en columnas L-N (centrado)
        monto = cargo.get('monto', '')
        ws[f'L{fila_cargo_fw}'] = f"$ {monto}" if monto else ''
        ws[f'L{fila_cargo_fw}'].font = Font(name='Arial', size=9, color='000000')
        ws[f'L{fila_cargo_fw}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'L{fila_cargo_fw}:N{fila_cargo_fw}')

        fila_cargo_fw += 1

    # Aplicar solo borde exterior a toda la seccion FreightWise
    ultima_fila_fw = fila_cargo_fw - 1
    thin_side = Side(style='thin', color='000000')
    no_side = Side(style=None)

    for row in range(fila_fw, ultima_fila_fw + 1):
        for col in range(1, 16):
            left = thin_side if col == 1 else no_side
            right = thin_side if col == 15 else no_side
            top = thin_side if row == fila_fw else no_side
            bottom = thin_side if row == ultima_fila_fw else no_side
            ws.cell(row=row, column=col).border = Border(left=left, right=right, top=top, bottom=bottom)

    # Notas FreightWise
    notas_fw = (datos.get('notas_freightwise') or '').strip()
    if notas_fw:
        fila_notas = ultima_fila_fw + 2

        # Titulo gris
        titulo_notas = 'Notas:' if idioma == 'es' else 'Notes:'
        ws[f'A{fila_notas}'] = titulo_notas
        ws[f'A{fila_notas}'].font = Font(name='Arial', size=11, bold=True, color='000000')
        ws[f'A{fila_notas}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'A{fila_notas}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{fila_notas}:O{fila_notas}')
        ws.row_dimensions[fila_notas].height = 25

        # Contenido morado con letras blancas
        fila_notas += 1
        ws[f'A{fila_notas}'] = notas_fw
        ws[f'A{fila_notas}'].font = Font(name='Arial', size=9, color='FFFFFF')
        ws[f'A{fila_notas}'].fill = PatternFill(start_color='4535A8', end_color='4535A8', fill_type='solid')
        ws[f'A{fila_notas}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'A{fila_notas}:O{fila_notas}')
        ws.row_dimensions[fila_notas].height = max(30, 15 * notas_fw.count('\n') + 15)

        # Borde exterior
        for row_n in range(fila_notas - 1, fila_notas + 1):
            for col in range(1, 16):
                left = thin_side if col == 1 else no_side
                right = thin_side if col == 15 else no_side
                top = thin_side if row_n == fila_notas - 1 else no_side
                bottom = thin_side if row_n == fila_notas else no_side
                ws.cell(row=row_n, column=col).border = Border(left=left, right=right, top=top, bottom=bottom)


def crear_cotizacion_excel(datos, idioma='ambos'):
    """Crea un archivo Excel segun el idioma seleccionado: 'es', 'en' o 'ambos'."""
    wb = Workbook()
    wb.remove(wb.active)

    if idioma in ('es', 'ambos'):
        ws_es = wb.create_sheet("Cotizacion Espanol")
        generar_hoja(ws_es, datos, idioma='es')

    if idioma in ('en', 'ambos'):
        ws_en = wb.create_sheet("Quotation English")
        generar_hoja(ws_en, datos, idioma='en')

    return wb


def guardar_cotizacion_bytes(datos, idioma='ambos'):
    """Genera la cotizacion y retorna los bytes del archivo."""
    wb = crear_cotizacion_excel(datos, idioma=idioma)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def guardar_cotizacion_archivo(datos, ruta_guardado=None):
    """Genera la cotizacion y la guarda en disco."""
    wb = crear_cotizacion_excel(datos)

    nombre_base = f"FreightWise_Cotizacion_{datos.get('origen', 'XXX')}-{datos.get('destino', 'XXX')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    if ruta_guardado:
        ruta_completa = os.path.join(ruta_guardado, nombre_base)
    else:
        ruta_completa = nombre_base

    wb.save(ruta_completa)

    return ruta_completa
