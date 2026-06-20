#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta las hojas del modulo Excel Online a un archivo .xlsx descargable,
preservando el formato por celda (negrita, color, alineacion, numero) que
ya soporta el editor."""

import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PURPLE = '4535A8'
_VALIGN_MAP = {'top': 'top', 'middle': 'center', 'bottom': 'bottom'}


def _cell_value(cell):
    if cell is None:
        return ''
    if isinstance(cell, dict):
        return cell.get('v', '') or ''
    return str(cell)


def _cell_fmt(cell):
    return cell if isinstance(cell, dict) else {}


def _numero(valor):
    # "+100" es una etiqueta de rango de kg, no un numero positivo literal:
    # no convertir, para no perder el signo "+".
    if isinstance(valor, str) and valor.strip().startswith('+'):
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _formato_numero(nf, dec):
    dec = dec if dec is not None else 2
    decimales = '.' + ('0' * dec) if dec else ''
    if nf == 'number':
        return f'0{decimales}'
    if nf == 'currency':
        return f'"$"#,##0{decimales}'
    if nf == 'accounting':
        return f'"$"#,##0{decimales};("$"#,##0{decimales})'
    if nf == 'percentage':
        return f'0{decimales}"%"'
    return 'General'


def _nombre_hoja_valido(nombre, usados):
    limpio = re.sub(r'[\[\]:\*\?/\\]', ' ', nombre or 'Hoja').strip()[:31] or 'Hoja'
    base, i = limpio, 2
    while limpio in usados:
        sufijo = f' ({i})'
        limpio = base[:31 - len(sufijo)] + sufijo
        i += 1
    usados.add(limpio)
    return limpio


def generar_excel_online_bytes(sheets):
    """sheets: lista de dicts con name/headers/rows, tal como ExcelSheet.to_dict()."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='B7AEDE')
    header_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    thin_gray = Side(style='thin', color='D9D9D9')
    bold_side = Side(style='medium', color='000000')

    nombres_usados = set()
    for sheet in sheets:
        ws = wb.create_sheet(_nombre_hoja_valido(sheet.get('name'), nombres_usados))
        headers = sheet.get('headers') or []
        rows = sheet.get('rows') or []

        for col_i, titulo in enumerate(headers, start=1):
            celda = ws.cell(row=1, column=col_i, value=titulo)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = header_border
        ws.row_dimensions[1].height = 20

        for row_i, fila in enumerate(rows, start=2):
            for col_i in range(1, len(headers) + 1):
                raw = fila[col_i - 1] if col_i - 1 < len(fila) else ''
                fmt = _cell_fmt(raw)
                texto = _cell_value(raw)
                numero = _numero(texto)

                celda = ws.cell(row=row_i, column=col_i, value=numero if numero is not None else texto)

                font_kwargs = {
                    'name': (fmt.get('ff') or 'Calibri').split(',')[0].strip().strip("'\""),
                    'size': fmt.get('fs') or 11,
                    'bold': bool(fmt.get('b')),
                    'italic': bool(fmt.get('i')),
                    'color': (fmt.get('fc') or '000000').lstrip('#'),
                }
                if fmt.get('u'):
                    font_kwargs['underline'] = 'single'
                if fmt.get('s'):
                    font_kwargs['strike'] = True
                celda.font = Font(**font_kwargs)

                if fmt.get('bg'):
                    color = fmt['bg'].lstrip('#')
                    celda.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                celda.alignment = Alignment(
                    horizontal=fmt.get('ha') or 'general',
                    vertical=_VALIGN_MAP.get(fmt.get('va'), 'center'),
                )

                if fmt.get('bd'):
                    celda.border = Border(top=bold_side, bottom=bold_side, left=bold_side, right=bold_side)
                else:
                    celda.border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

                if numero is not None and fmt.get('nf'):
                    celda.number_format = _formato_numero(fmt.get('nf'), fmt.get('dec'))

        for col_i, titulo in enumerate(headers, start=1):
            ancho = max(12, min(40, len(str(titulo or '')) + 4))
            ws.column_dimensions[get_column_letter(col_i)].width = ancho
        if rows:
            ws.freeze_panes = 'A2'

    if not wb.sheetnames:
        wb.create_sheet('Hoja1')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
