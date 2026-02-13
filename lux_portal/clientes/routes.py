#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rutas del modulo Clientes
"""

from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session
from lux_portal.clientes import clientes_bp
from lux_portal.clientes.models import (
    Cliente, ClientePlanilla, RUTA_TEMPLATE,
    CustomerAssociateForm, ShippingInstructionsForm, FreightQuoteForm, PaymentInfoForm
)
from lux_portal.extensions import db
from lux_portal.auth.decorators import login_required
import json
import pandas as pd
from io import BytesIO
import os


@clientes_bp.route('/')
@login_required
def dashboard():
    """Lista de clientes con filtro"""
    busqueda = request.args.get('q', '').strip()

    query = Cliente.query.filter_by(activo=True)

    if busqueda:
        query = query.filter(
            db.or_(
                Cliente.codigo.ilike(f'%{busqueda}%'),
                Cliente.nombre.ilike(f'%{busqueda}%'),
                Cliente.ruta_principal.ilike(f'%{busqueda}%')
            )
        )

    clientes = query.order_by(Cliente.fecha_actualizacion.desc()).all()

    return render_template(
        'clientes/dashboard.html',
        clientes=clientes,
        busqueda=busqueda
    )


@clientes_bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_cliente():
    """Crear nuevo cliente"""
    if request.method == 'POST':
        codigo = request.form.get('codigo', '').strip().upper()
        nombre = request.form.get('nombre', '').strip()
        ruta_principal = request.form.get('ruta_principal', '').strip().upper()

        if not codigo:
            flash('El codigo del cliente es requerido', 'danger')
            return redirect(url_for('clientes.nuevo_cliente'))

        # Verificar si ya existe (activo)
        existe_activo = Cliente.query.filter_by(codigo=codigo, activo=True).first()
        if existe_activo:
            flash(f'Ya existe un cliente con el codigo {codigo}', 'danger')
            return redirect(url_for('clientes.nuevo_cliente'))

        # Verificar si existe pero inactivo (reactivar)
        existe_inactivo = Cliente.query.filter_by(codigo=codigo, activo=False).first()
        if existe_inactivo:
            existe_inactivo.activo = True
            existe_inactivo.nombre = nombre or codigo
            existe_inactivo.ruta_principal = ruta_principal
            db.session.commit()
            flash(f'Cliente {codigo} reactivado exitosamente', 'success')
            return redirect(url_for('clientes.ver_planilla', id=existe_inactivo.id))

        # Crear cliente nuevo
        cliente = Cliente(
            codigo=codigo,
            nombre=nombre or codigo,
            ruta_principal=ruta_principal
        )
        db.session.add(cliente)
        db.session.flush()

        # Crear planilla vacia
        planilla = ClientePlanilla(
            cliente_id=cliente.id,
            rutas_data=json.dumps([RUTA_TEMPLATE.copy()])
        )
        db.session.add(planilla)
        db.session.commit()

        flash(f'Cliente {codigo} creado exitosamente', 'success')
        return redirect(url_for('clientes.ver_planilla', id=cliente.id))

    return render_template('clientes/nuevo.html')


@clientes_bp.route('/<int:id>')
@login_required
def ver_planilla(id):
    """Ver y editar planilla del cliente"""
    cliente = Cliente.query.get_or_404(id)

    # Crear planilla si no existe
    if not cliente.planilla:
        planilla = ClientePlanilla(
            cliente_id=cliente.id,
            rutas_data=json.dumps([RUTA_TEMPLATE.copy()])
        )
        db.session.add(planilla)
        db.session.commit()

    return render_template(
        'clientes/planilla.html',
        cliente=cliente,
        planilla=cliente.planilla,
        rutas=cliente.planilla.get_rutas()
    )


@clientes_bp.route('/api/cliente/<int:id>', methods=['PUT'])
@login_required
def actualizar_cliente(id):
    """API para actualizar datos basicos del cliente"""
    cliente = Cliente.query.get_or_404(id)
    data = request.get_json()

    if 'nombre' in data:
        cliente.nombre = data['nombre']
    if 'ruta_principal' in data:
        cliente.ruta_principal = data['ruta_principal'].upper()

    db.session.commit()
    return jsonify({'success': True})


@clientes_bp.route('/api/cliente/<int:id>', methods=['DELETE'])
@login_required
def eliminar_cliente(id):
    """API para eliminar cliente (soft delete)"""
    try:
        cliente = Cliente.query.get_or_404(id)
        cliente.activo = False
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@clientes_bp.route('/api/planilla/<int:id>/rutas', methods=['POST'])
@login_required
def guardar_rutas(id):
    """API para guardar datos de rutas"""
    cliente = Cliente.query.get_or_404(id)
    data = request.get_json()

    if not cliente.planilla:
        planilla = ClientePlanilla(cliente_id=cliente.id)
        db.session.add(planilla)
        db.session.flush()

    cliente.planilla.set_rutas(data.get('rutas', []))
    db.session.commit()

    return jsonify({'success': True})


@clientes_bp.route('/api/planilla/<int:id>/calculos', methods=['POST'])
@login_required
def guardar_calculos(id):
    """API para guardar datos de calculos"""
    cliente = Cliente.query.get_or_404(id)
    data = request.get_json()

    if not cliente.planilla:
        planilla = ClientePlanilla(cliente_id=cliente.id)
        db.session.add(planilla)
        db.session.flush()

    cliente.planilla.set_calculos(data.get('calculos', {}))
    db.session.commit()

    return jsonify({'success': True})


@clientes_bp.route('/api/planilla/<int:id>/notas', methods=['POST'])
@login_required
def guardar_notas(id):
    """API para guardar notas"""
    cliente = Cliente.query.get_or_404(id)
    data = request.get_json()

    if not cliente.planilla:
        planilla = ClientePlanilla(cliente_id=cliente.id)
        db.session.add(planilla)
        db.session.flush()

    cliente.planilla.notas = data.get('notas', '')
    db.session.commit()

    return jsonify({'success': True})


@clientes_bp.route('/api/planilla/<int:id>/banco', methods=['POST'])
@login_required
def guardar_banco(id):
    """API para guardar info bancaria"""
    try:
        cliente = Cliente.query.get_or_404(id)
        data = request.get_json()

        if not cliente.planilla:
            planilla = ClientePlanilla(cliente_id=cliente.id)
            db.session.add(planilla)
            db.session.flush()

        if 'banco' in data:
            cliente.planilla.banco = data['banco']
        if 'swift' in data:
            cliente.planilla.swift = data['swift']
        if 'cuenta' in data:
            cliente.planilla.cuenta = data['cuenta']
        if 'empresa' in data:
            cliente.planilla.empresa = data['empresa']
        if 'tax_id' in data:
            cliente.planilla.tax_id = data['tax_id']
        if 'direccion' in data:
            cliente.planilla.direccion = data['direccion']

        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@clientes_bp.route('/api/planilla/<int:id>/todo', methods=['POST'])
@login_required
def guardar_todo(id):
    """API para guardar toda la planilla de una vez"""
    try:
        cliente = Cliente.query.get_or_404(id)
        data = request.get_json()

        if not data:
            return jsonify({'success': False, 'error': 'No se recibieron datos'}), 400

        if not cliente.planilla:
            planilla = ClientePlanilla(cliente_id=cliente.id)
            db.session.add(planilla)
            db.session.flush()

        planilla = cliente.planilla

        if 'rutas' in data:
            planilla.set_rutas(data['rutas'])
        if 'notas' in data:
            planilla.notas = data['notas']

        # Actualizar fecha del cliente
        from datetime import datetime
        cliente.fecha_actualizacion = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True, 'mensaje': 'Planilla guardada'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# RUTAS DE FORMULARIOS
# ============================================================================

@clientes_bp.route('/formularios')
@login_required
def formularios():
    """Seleccion de tipo de formulario"""
    # Obtener formularios existentes
    customer_forms = CustomerAssociateForm.query.order_by(
        CustomerAssociateForm.fecha_actualizacion.desc()
    ).limit(10).all()

    shipping_forms = ShippingInstructionsForm.query.order_by(
        ShippingInstructionsForm.fecha_actualizacion.desc()
    ).limit(10).all()

    payment_forms = PaymentInfoForm.query.order_by(
        PaymentInfoForm.fecha_actualizacion.desc()
    ).limit(10).all()

    return render_template(
        'clientes/formularios.html',
        customer_forms=customer_forms,
        shipping_forms=shipping_forms,
        payment_forms=payment_forms
    )


@clientes_bp.route('/formularios/customer-associate', methods=['GET', 'POST'])
@clientes_bp.route('/formularios/customer-associate/<int:id>', methods=['GET', 'POST'])
@login_required
def customer_associate_form(id=None):
    """Formulario Customer Associate - manual o edicion"""
    form_data = None

    if id:
        form_data = CustomerAssociateForm.query.get_or_404(id)

    if request.method == 'POST':
        try:
            if id:
                form = CustomerAssociateForm.query.get_or_404(id)
            else:
                form = CustomerAssociateForm()

            # General Information
            form.legal_name = request.form.get('legal_name', '').strip()
            form.tax_id = request.form.get('tax_id', '').strip()
            form.company_type = request.form.get('company_type', '').strip()
            form.address = request.form.get('address', '').strip()
            form.city_country = request.form.get('city_country', '').strip()
            form.telephone = request.form.get('telephone', '').strip()
            emails = [e.strip() for e in request.form.getlist('email[]') if e.strip()]
            form.email = ', '.join(emails)
            form.website = request.form.get('website', '').strip()
            form.business_activity = request.form.get('business_activity', '').strip()
            form.annual_revenue = request.form.get('annual_revenue', '').strip()

            # Legal Representative
            form.legal_rep_name = request.form.get('legal_rep_name', '').strip()
            form.legal_rep_id = request.form.get('legal_rep_id', '').strip()
            form.legal_rep_telephone = request.form.get('legal_rep_telephone', '').strip()
            form.legal_rep_email = request.form.get('legal_rep_email', '').strip()

            # Shareholders
            shareholders = []
            sh_names = request.form.getlist('shareholder_name[]')
            sh_ids = request.form.getlist('shareholder_id[]')
            sh_nationalities = request.form.getlist('shareholder_nationality[]')
            sh_percentages = request.form.getlist('shareholder_percentage[]')
            for i, name in enumerate(sh_names):
                if name.strip():
                    shareholders.append({
                        'name': name.strip(),
                        'id_number': sh_ids[i].strip() if i < len(sh_ids) else '',
                        'nationality': sh_nationalities[i].strip() if i < len(sh_nationalities) else '',
                        'percentage': sh_percentages[i].strip() if i < len(sh_percentages) else '',
                    })
            form.set_shareholders(shareholders)

            # Financial Contact (name + multiple emails + multiple phones)
            fc_emails = [e.strip() for e in request.form.getlist('fc_email[]') if e.strip()]
            fc_phones = [p.strip() for p in request.form.getlist('fc_phone[]') if p.strip()]
            form.set_financial_contact({
                'name': request.form.get('fc_name', '').strip(),
                'emails': fc_emails,
                'phones': fc_phones,
            })
            form.bank_name = request.form.get('bank_name', '').strip()
            form.bank_account = request.form.get('bank_account', '').strip()
            form.bank_address = request.form.get('bank_address', '').strip()
            form.swift_aba = request.form.get('swift_aba', '').strip()

            # Ethics
            ethics = {
                'q1': request.form.get('ethics_q1', ''),
                'q2': request.form.get('ethics_q2', ''),
                'q3': request.form.get('ethics_q3', ''),
                'q4': request.form.get('ethics_q4', ''),
                'q5': request.form.get('ethics_q5', ''),
                'q6': request.form.get('ethics_q6', ''),
                'q7': request.form.get('ethics_q7', '')
            }
            form.set_ethics(ethics)

            form.estado = 'completo'

            if not id:
                db.session.add(form)

            db.session.commit()
            flash('Formulario guardado exitosamente', 'success')
            return redirect(url_for('clientes.formularios'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar: {str(e)}', 'danger')

    return render_template(
        'clientes/form_customer_associate.html',
        form_data=form_data
    )


@clientes_bp.route('/formularios/shipping-instructions', methods=['GET', 'POST'])
@clientes_bp.route('/formularios/shipping-instructions/<int:id>', methods=['GET', 'POST'])
@login_required
def shipping_instructions_form(id=None):
    """Formulario Shipping Instructions - manual o edicion"""
    form_data = None

    if id:
        form_data = ShippingInstructionsForm.query.get_or_404(id)

    if request.method == 'POST':
        try:
            if id:
                form = ShippingInstructionsForm.query.get_or_404(id)
            else:
                form = ShippingInstructionsForm()

            # Contact Information
            form.contact_full_name = request.form.get('contact_full_name', '').strip()
            form.contact_job_title = request.form.get('contact_job_title', '').strip()
            form.contact_email = request.form.get('contact_email', '').strip()
            form.contact_phone = request.form.get('contact_phone', '').strip()
            form.contact_whatsapp = request.form.get('contact_whatsapp', '').strip()

            # Billing
            form.billing_address = request.form.get('billing_address', '').strip()
            form.billing_emails = request.form.get('billing_emails', '').strip()

            # Additional
            form.uses_customs_broker = request.form.get('uses_customs_broker', '').strip()
            form.customs_broker_info = request.form.get('customs_broker_info', '').strip()
            form.customs_broker_emails = request.form.get('customs_broker_emails', '').strip()
            form.api_required = request.form.get('api_required', '').strip()

            # Documents
            documents = {
                'invoice': 'invoice' in request.form,
                'coa': 'coa' in request.form,
                'phyto': 'phyto' in request.form,
                'phyto_destination': 'phyto_destination' in request.form,
                'master_invoice': 'master_invoice' in request.form,
                'other': request.form.get('documents_other', '')
            }
            form.set_documents(documents)

            # Suppliers
            suppliers = []
            sup_names = request.form.getlist('supplier_name[]')
            sup_emails = request.form.getlist('supplier_email[]')
            for i, name in enumerate(sup_names):
                if name.strip():
                    suppliers.append({
                        'name': name.strip(),
                        'email': sup_emails[i].strip() if i < len(sup_emails) else '',
                    })
            form.set_suppliers(suppliers)

            form.estado = 'completo'

            if not id:
                db.session.add(form)

            db.session.commit()
            flash('Formulario guardado exitosamente', 'success')
            return redirect(url_for('clientes.formularios'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar: {str(e)}', 'danger')

    return render_template(
        'clientes/form_shipping_instructions.html',
        form_data=form_data
    )


@clientes_bp.route('/formularios/upload/<form_type>', methods=['GET', 'POST'])
@login_required
def upload_form(form_type):
    """Subir Excel o PDF para auto-llenar formulario"""
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No se selecciono ningun archivo', 'danger')
            return redirect(request.url)

        file = request.files['excel_file']
        if file.filename == '':
            flash('No se selecciono ningun archivo', 'danger')
            return redirect(request.url)

        fname = file.filename.lower()
        is_pdf = fname.endswith('.pdf')
        is_excel = fname.endswith(('.xlsx', '.xls'))

        if not is_pdf and not is_excel:
            flash('El archivo debe ser Excel (.xlsx, .xls) o PDF (.pdf)', 'danger')
            return redirect(request.url)

        try:
            if form_type == 'customer-associate':
                if is_pdf:
                    form_data = parse_customer_associate_pdf(file)
                else:
                    df = pd.read_excel(file, header=None)
                    form_data = parse_customer_associate_excel(df)
                # Save directly to DB
                form = CustomerAssociateForm()
                for key, value in form_data.items():
                    if key == 'shareholders':
                        form.set_shareholders(value)
                    elif key == 'ethics':
                        form.set_ethics(value)
                    elif hasattr(form, key):
                        setattr(form, key, value)
                form.estado = 'borrador'
                db.session.add(form)
                db.session.commit()
                flash('Datos importados exitosamente. Revisa y guarda.', 'success')
                return redirect(url_for('clientes.customer_associate_form', id=form.id))

            elif form_type == 'shipping-instructions':
                if is_pdf:
                    form_data = parse_shipping_pdf(file)
                else:
                    df = pd.read_excel(file, header=None)
                    form_data = parse_shipping_instructions_excel(df)
                form = ShippingInstructionsForm()
                for key, value in form_data.items():
                    if key == 'documents':
                        form.set_documents(value)
                    elif key == 'suppliers':
                        form.set_suppliers(value)
                    elif hasattr(form, key):
                        setattr(form, key, value)
                form.estado = 'borrador'
                db.session.add(form)
                db.session.commit()
                flash('Datos importados exitosamente. Revisa y guarda.', 'success')
                return redirect(url_for('clientes.shipping_instructions_form', id=form.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el archivo: {str(e)}', 'danger')
            return redirect(request.url)

    return render_template(
        'clientes/upload_form.html',
        form_type=form_type
    )


@clientes_bp.route('/formularios/customer-associate/<int:id>/descargar')
@login_required
def descargar_customer_associate(id):
    """Descargar Excel con datos del Customer Associate Form"""
    form = CustomerAssociateForm.query.get_or_404(id)

    try:
        # Usar plantilla original como base
        template_path = r'C:\Users\nicol\Downloads\FW CUSTOMER ASSOCIATE FORMAT.xlsx'
        output = BytesIO()

        if os.path.exists(template_path):
            from openpyxl import load_workbook
            wb = load_workbook(template_path)
            ws = wb.active

            # Mapear datos a celdas (columna C)
            mappings = {
                10: form.legal_name,
                11: form.tax_id,
                12: form.company_type,
                13: form.address,
                14: form.city_country,
                15: form.telephone,
                16: form.email,
                17: form.website,
                18: form.business_activity,
                19: form.annual_revenue,
                21: form.legal_rep_name,
                22: form.legal_rep_id,
                23: form.legal_rep_telephone,
                24: form.legal_rep_email,
                33: form.financial_contact,
                34: form.bank_name,
                35: form.bank_account,
                36: form.bank_address,
                37: form.swift_aba,
            }

            for row, value in mappings.items():
                if value:
                    ws.cell(row=row, column=3, value=value)

            # Agregar shareholders
            shareholders = form.get_shareholders()
            for i, sh in enumerate(shareholders[:2]):  # Max 2 rows
                ws.cell(row=29 + i, column=2, value=sh.get('name', ''))
                ws.cell(row=29 + i, column=3, value=sh.get('id_number', ''))

            wb.save(output)
        else:
            # Fallback: crear Excel simple
            data = {
                'Campo': ['Legal Name', 'Tax ID', 'Company Type', 'Address', 'City-Country',
                         'Telephone', 'Email', 'Website', 'Business Activity', 'Annual Revenue'],
                'Valor': [form.legal_name, form.tax_id, form.company_type, form.address,
                         form.city_country, form.telephone, form.email, form.website,
                         form.business_activity, form.annual_revenue]
            }
            df_out = pd.DataFrame(data)
            df_out.to_excel(output, index=False)

        output.seek(0)

        filename = f'CustomerAssociate_{form.id}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'danger')
        return redirect(url_for('clientes.formularios'))


@clientes_bp.route('/formularios/shipping-instructions/<int:id>/descargar')
@login_required
def descargar_shipping_instructions(id):
    """Descargar Excel con datos del Shipping Instructions Form"""
    form = ShippingInstructionsForm.query.get_or_404(id)

    try:
        template_path = r'C:\Users\nicol\Downloads\FWCustomer Shipping Instructions.xlsx'
        output = BytesIO()

        if os.path.exists(template_path):
            from openpyxl import load_workbook
            wb = load_workbook(template_path)
            ws = wb.active

            # Mapear datos a celdas (columna B)
            mappings = {
                9: form.contact_full_name,
                10: form.contact_job_title,
                11: form.contact_email,
                12: form.contact_phone,
                13: form.contact_whatsapp,
                16: form.billing_address,
                17: form.billing_emails,
                24: form.uses_customs_broker,
                25: form.customs_broker_info,
                26: form.customs_broker_emails,
            }

            for row, value in mappings.items():
                if value:
                    ws.cell(row=row, column=2, value=value)

            # Documents
            docs = form.get_documents()
            if docs.get('invoice'):
                ws.cell(row=30, column=2, value='x')
            if docs.get('coa'):
                ws.cell(row=31, column=2, value='x')
            if docs.get('phyto'):
                ws.cell(row=32, column=2, value='x')

            wb.save(output)
        else:
            # Fallback
            data = {
                'Campo': ['Contact Name', 'Job Title', 'Email', 'Phone', 'WhatsApp',
                         'Billing Address', 'Billing Emails'],
                'Valor': [form.contact_full_name, form.contact_job_title, form.contact_email,
                         form.contact_phone, form.contact_whatsapp, form.billing_address,
                         form.billing_emails]
            }
            df_out = pd.DataFrame(data)
            df_out.to_excel(output, index=False)

        output.seek(0)

        filename = f'ShippingInstructions_{form.id}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'danger')
        return redirect(url_for('clientes.formularios'))


@clientes_bp.route('/formularios/customer-associate/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_customer_associate(id):
    """Eliminar formulario Customer Associate"""
    try:
        form = CustomerAssociateForm.query.get_or_404(id)
        db.session.delete(form)
        db.session.commit()
        flash('Formulario eliminado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('clientes.formularios'))


@clientes_bp.route('/formularios/shipping-instructions/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_shipping_instructions(id):
    """Eliminar formulario Shipping Instructions"""
    try:
        form = ShippingInstructionsForm.query.get_or_404(id)
        db.session.delete(form)
        db.session.commit()
        flash('Formulario eliminado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('clientes.formularios'))


# ============================================================================
# RUTAS DE PAYMENT INFO FORM
# ============================================================================

@clientes_bp.route('/formularios/payment-info', methods=['GET', 'POST'])
@clientes_bp.route('/formularios/payment-info/<int:id>', methods=['GET', 'POST'])
@login_required
def payment_info_form(id=None):
    """Formulario Invoice"""
    form_data = None
    if id:
        form_data = PaymentInfoForm.query.get_or_404(id)

    if request.method == 'POST':
        try:
            form = PaymentInfoForm.query.get_or_404(id) if id else PaymentInfoForm()

            # Header
            for f in ['invoice_date', 'expiration_date', 'invoice_number',
                       'customer_id_code', 'date_from', 'date_to']:
                setattr(form, f, request.form.get(f, '').strip())

            # Customer
            form.customer_name = request.form.get('customer_name', '').strip()
            form.customer_address = request.form.get('customer_address', '').strip()

            # Shipment
            for f in ['peso', 'moneda', 'aerolinea', 'tarifa', 'origen', 'destino', 'tarifa_iva']:
                setattr(form, f, request.form.get(f, '').strip())

            # Line Items
            li_dates = request.form.getlist('li_date[]')
            li_units = request.form.getlist('li_units[]')
            li_descs = request.form.getlist('li_desc[]')
            li_rates = request.form.getlist('li_rate[]')
            li_amounts = request.form.getlist('li_amount[]')
            li_totals = request.form.getlist('li_total[]')
            items = []
            for i in range(len(li_descs)):
                if li_descs[i].strip() or (i < len(li_rates) and li_rates[i].strip()):
                    items.append({
                        'date': li_dates[i].strip() if i < len(li_dates) else '',
                        'units': li_units[i].strip() if i < len(li_units) else '',
                        'description': li_descs[i].strip(),
                        'rate': li_rates[i].strip() if i < len(li_rates) else '',
                        'amount': li_amounts[i].strip() if i < len(li_amounts) else '',
                        'total': li_totals[i].strip() if i < len(li_totals) else '',
                    })
            form.set_line_items(items)

            # Terms
            for f in ['full_count', 'pieces_count', 'awb', 'gross_weight']:
                setattr(form, f, request.form.get(f, '').strip())

            # Bank
            for f in ['fw_bank', 'fw_swift', 'fw_account', 'fw_tax_id', 'fw_company']:
                setattr(form, f, request.form.get(f, '').strip())

            # Totals
            for f in ['subtotal', 'taxable', 'tax_rate_vat', 'tax', 'other_charges',
                       'insurance', 'legal_consular', 'inspection_cert', 'other_specify',
                       'total', 'currency']:
                setattr(form, f, request.form.get(f, '').strip())

            form.reason_for_export = request.form.get('reason_for_export', '').strip()
            form.label = form.invoice_number or form.customer_name or ''
            form.estado = 'completo'

            if not id:
                db.session.add(form)
            db.session.commit()
            flash('Invoice guardada exitosamente', 'success')
            return redirect(url_for('clientes.payment_info_form', id=form.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar: {str(e)}', 'danger')

    return render_template('clientes/form_payment_info.html', form_data=form_data)


@clientes_bp.route('/formularios/payment-info/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_payment_info(id):
    try:
        form = PaymentInfoForm.query.get_or_404(id)
        db.session.delete(form)
        db.session.commit()
        flash('Registro eliminado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('clientes.formularios'))


@clientes_bp.route('/formularios/payment-info/<int:id>/pdf')
@login_required
def descargar_invoice_pdf(id):
    """Genera PDF de la invoice"""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    form = PaymentInfoForm.query.get_or_404(id)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

    BLUE = colors.HexColor('#4472C4')
    LIGHT_BLUE = colors.HexColor('#D6E4F0')
    BLACK = colors.black
    WHITE = colors.white

    s_normal = ParagraphStyle('sn', fontSize=8, leading=10)
    s_bold = ParagraphStyle('sb', fontSize=8, leading=10, fontName='Helvetica-Bold')
    s_center = ParagraphStyle('sc', fontSize=8, leading=10, alignment=TA_CENTER)
    s_center_bold = ParagraphStyle('scb', fontSize=8, leading=10, alignment=TA_CENTER, fontName='Helvetica-Bold')
    s_right = ParagraphStyle('sr', fontSize=8, leading=10, alignment=TA_RIGHT)
    s_right_bold = ParagraphStyle('srb', fontSize=8, leading=10, alignment=TA_RIGHT, fontName='Helvetica-Bold')
    s_small = ParagraphStyle('ss', fontSize=7, leading=9)
    s_title_white = ParagraphStyle('stw', fontSize=10, leading=12, fontName='Helvetica-Bold',
                                    textColor=WHITE, alignment=TA_CENTER)
    P = Paragraph

    # ======== HEADER: Company + Invoice info ========
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'images', 'logo.png')

    # Left: Company info
    company_info = [
        [P('<b>FreightWise Forwarding S.A.</b>', ParagraphStyle('ci', fontSize=11, leading=13, fontName='Helvetica-Bold', alignment=TA_CENTER))],
        [P(f'RUC: {form.fw_tax_id}', s_center)],
        [P('', s_center)],
        [P('Tababela Cargo Center', s_center)],
        [P('Quito - Ecuador', s_center)],
        [P('www.freigh-twise.com', s_center)],
    ]
    company_table = Table(company_info, colWidths=[65*mm])
    company_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BLUE),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Right: Invoice details
    inv_data = [
        ['', '', P('<b>INVOICE</b>', ParagraphStyle('inv', fontSize=14, fontName='Helvetica-Bold', alignment=TA_RIGHT))],
        [P('Date', s_right), '', P(form.invoice_date or '', s_right_bold)],
        [P('Expiration Date', s_right), '', P(form.expiration_date or '', s_right_bold)],
        [P('Invoice #', s_right), '', P(form.invoice_number or '', s_right_bold)],
        [P('Customer ID', s_right), '', P(form.customer_id_code or '', s_right_bold)],
        ['', '', ''],
        [P('From:', s_right), P(form.date_from or '', s_center), P('to:  ' + (form.date_to or ''), s_normal)],
    ]
    inv_table = Table(inv_data, colWidths=[30*mm, 25*mm, 40*mm])
    inv_table.setStyle(TableStyle([
        ('GRID', (0, 1), (-1, 4), 0.5, colors.grey),
        ('GRID', (0, 6), (-1, 6), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Customer box
    cust_data = [
        [P('<b>CUSTOMER:</b>', s_bold)],
        [P(form.customer_name or '', s_bold)],
        [P(form.customer_address or '', s_small)],
    ]
    cust_table = Table(cust_data, colWidths=[65*mm])
    cust_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))

    # Combine header
    header_data = [[company_table, '', inv_table], ['', '', cust_table]]
    header = Table(header_data, colWidths=[70*mm, 20*mm, 95*mm])
    header.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.append(header)
    elements.append(Spacer(1, 5*mm))

    # ======== SHIPMENT INFO ========
    ship_data = [[
        P(f'Peso:', s_small), P(f'<b>{form.peso or ""}</b>', s_small),
        P(f'Moneda:', s_small), P(f'<b>{form.moneda or "USD"}</b>', s_small),
        '', '',
    ], [
        P(f'Tarifa:', s_small), P(f'<b>{form.tarifa or ""}</b>', s_small),
        P(f'Aerolinea:', s_small), P(f'<b>{form.aerolinea or ""}</b>', s_small),
        '', '',
    ], [
        P(f'Origen:', s_small), P(f'<b>{form.origen or "UIO"}</b>', s_small),
        P(f'Destino', s_small), P(f'<b>{form.destino or ""}</b>', s_small),
        P(f'Tarifa IVA:', s_small), P(f'<b>{form.tarifa_iva or "0%"}</b>', s_small),
    ]]
    ship_table = Table(ship_data, colWidths=[18*mm, 20*mm, 20*mm, 30*mm, 22*mm, 20*mm])
    ship_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements.append(ship_table)
    elements.append(Spacer(1, 3*mm))

    # ======== LINE ITEMS TABLE ========
    items = form.get_line_items()
    li_header = [P('Date', s_center_bold), P('UNITS', s_center_bold), '',
                 P('DESCRIPTION', s_center_bold), '',
                 P('RATE<br/>PRICE', s_center_bold), P('AMOUNT', s_center_bold),
                 P('TOTAL<br/>AMOUNT', s_center_bold)]
    li_rows = [li_header]
    for item in items:
        li_rows.append([
            P(item.get('date', ''), s_center), P(item.get('units', ''), s_center), '',
            P(item.get('description', ''), s_normal), '',
            P(item.get('rate', ''), s_right), P(item.get('amount', ''), s_right),
            P(item.get('total', ''), s_right),
        ])
    # Fill empty rows to match template (at least 14 rows)
    while len(li_rows) < 15:
        li_rows.append(['', '', '', '', '', '', '', P('-', s_right)])

    li_col_w = [18*mm, 14*mm, 5*mm, 45*mm, 10*mm, 18*mm, 25*mm, 25*mm]
    li_table = Table(li_rows, colWidths=li_col_w)
    li_style = [
        ('BACKGROUND', (0, 0), (-1, 0), BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('GRID', (0, 0), (-1, 0), 0.5, BLACK),
        ('LINEBELOW', (0, -1), (-1, -1), 0.5, BLACK),
        ('LINEBEFORE', (0, 0), (0, -1), 0.5, BLACK),
        ('LINEAFTER', (-1, 0), (-1, -1), 0.5, BLACK),
        ('LINEAFTER', (-2, 0), (-2, -1), 0.5, BLACK),
        ('SPAN', (2, 0), (4, 0)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]
    # Span description columns for data rows
    for r in range(1, len(li_rows)):
        li_style.append(('SPAN', (2, r), (4, r)))
    li_table.setStyle(TableStyle(li_style))

    # ======== TOTALS (right side of line items) ========
    totals_data = [
        [P('Subtotal', s_right), P('$', s_center), P(form.subtotal or '-', s_right)],
        [P('Taxable', s_right), P('$', s_center), P(form.taxable or '-', s_right)],
        [P('Tax rate VAT', s_right), '', P(form.tax_rate_vat or '0.000%', s_right)],
        [P('Tax', s_right), P('$', s_center), P(form.tax or '-', s_right)],
        [P('Other charges', s_right), P('$', s_center), P(form.other_charges or '-', s_right)],
        [P('Insurance', s_right), P('$', s_center), P(form.insurance or '-', s_right)],
        [P('Legal/Consular', s_right), P('$', s_center), P(form.legal_consular or '-', s_right)],
        [P('Inspection/Cert.', s_right), P('$', s_center), P(form.inspection_cert or '-', s_right)],
        [P('Other (specify)', s_right), P('$', s_center), P(form.other_specify or '-', s_right)],
        [P('<b>TOTAL</b>', s_right_bold), P('<b>$</b>', s_center_bold), P(f'<b>{form.total or "-"}</b>', s_right_bold)],
        [P('Currency', s_right), '', P(form.currency or 'USD', s_right)],
    ]
    totals_table = Table(totals_data, colWidths=[28*mm, 8*mm, 20*mm])
    totals_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))

    # Combine line items + totals side by side
    main_data = [[li_table, totals_table]]
    main_table = Table(main_data, colWidths=[160*mm, 56*mm])
    main_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.append(main_table)
    elements.append(Spacer(1, 3*mm))

    # ======== TERMS + BANK ========
    terms_bank = [
        [P('<b>TERMS OF SALE AND OTHER COMMENTS</b>', s_small), '', '', '', '', ''],
        [P('FULL:', s_small), P(form.full_count or '0', s_small), '',
         P('AWB:', s_small), '', P(form.awb or '', s_small)],
        [P('PIECES:', s_small), P(form.pieces_count or '0', s_small), '',
         P('GROSS:', s_small), '', P(form.gross_weight or '', s_small)],
        ['', '', '', '', '', ''],
        [P('BANK', s_bold), P(form.fw_bank or '', s_small), '', '', '', ''],
        [P('SWIFT', s_bold), P(form.fw_swift or '', s_small), '', '', '', ''],
        [P('ACCOUNT', s_bold), P(form.fw_account or '', s_small), '', '', '', ''],
        [P('TAX ID', s_bold), P(form.fw_tax_id or '', s_small), '', '', '', ''],
        [P('COMPANY', s_bold), P(form.fw_company or '', s_small), '', '', '', ''],
    ]
    tb_table = Table(terms_bank, colWidths=[20*mm, 35*mm, 10*mm, 15*mm, 5*mm, 30*mm])
    tb_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BLUE),
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, BLACK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(tb_table)
    elements.append(Spacer(1, 5*mm))

    # ======== ADDITIONAL DETAILS ========
    add_data = [
        [P('<b>ADDITIONAL DETAILS</b>', s_small), '', '', ''],
        ['', '', '', ''],
        [P('Reason for Export:', s_small), '', P(form.reason_for_export or 'Fresh Flowers', s_center_bold), ''],
        ['', '', '', ''],
        [P('I certify the above to be true and correct to the best of my knowledge.', s_small), '', '', ''],
        ['', '', P(form.invoice_date or '', s_center), ''],
        ['', '', P('Date', s_center), ''],
    ]
    add_table = Table(add_data, colWidths=[60*mm, 15*mm, 50*mm, 40*mm])
    add_table.setStyle(TableStyle([
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BLUE),
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, BLACK),
        ('BOX', (2, 2), (2, 2), 0.5, BLACK),
        ('LINEABOVE', (2, 5), (2, 5), 0.5, BLACK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('SPAN', (0, 4), (-1, 4)),
    ]))
    elements.append(add_table)

    # Page number
    elements.append(Spacer(1, 5*mm))
    elements.append(P('Page 1 of 1', s_center))

    doc.build(elements)
    output.seek(0)
    inv_num = (form.invoice_number or str(form.id)).replace(' ', '_')
    filename = f'Invoice_{inv_num}.pdf'
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ============================================================================
# RUTAS DE FREIGHT QUOTE FORM
# ============================================================================

@clientes_bp.route('/formularios/freight-quote', methods=['GET', 'POST'])
@clientes_bp.route('/formularios/freight-quote/<int:id>', methods=['GET', 'POST'])
@login_required
def freight_quote_form(id=None):
    """Formulario Freight Quote - manual o edicion"""
    form_data = None
    uploaded_data = session.pop('uploaded_form_data', None)
    uploaded_type = session.pop('uploaded_form_type', None)

    if id:
        form_data = FreightQuoteForm.query.get_or_404(id)
    elif uploaded_data and uploaded_type == 'freight-quote':
        # Usar datos del Excel subido
        form_data = type('FormData', (), uploaded_data)()
        form_data.id = None

    if request.method == 'POST':
        try:
            if id:
                form = FreightQuoteForm.query.get_or_404(id)
            else:
                form = FreightQuoteForm()

            # Ruta y pesos
            form.ruta = request.form.get('ruta', '').strip().upper()
            form.net_weight = float(request.form.get('net_weight', 0) or 0)
            form.volume_weight = float(request.form.get('volume_weight', 0) or 0)
            form.rate = float(request.form.get('rate', 0) or 0)

            # Cargos
            form.due_agent = float(request.form.get('due_agent', 0) or 0)
            form.due_carrier = float(request.form.get('due_carrier', 0) or 0)
            form.certificate = float(request.form.get('certificate', 0) or 0)
            form.phyto = float(request.form.get('phyto', 0) or 0)

            # Calcular totales
            form.calculate_totals()

            # Notas
            form.notes = request.form.get('notes', '').strip()

            # Datos bancarios
            form.bank_name = request.form.get('bank_name', 'Banco Pichincha').strip()
            form.bank_swift = request.form.get('bank_swift', 'PICHECEQXXX').strip()
            form.bank_account = request.form.get('bank_account', '2100339784').strip()
            form.company_name = request.form.get('company_name', 'FREIGHTWISE FORWARDING S.A').strip()
            form.tax_id = request.form.get('tax_id', '1793230131001').strip()
            form.company_address = request.form.get('company_address', 'Pedro de Alfaro y Francisco Ruiz').strip()

            form.estado = 'completo'

            if not id:
                db.session.add(form)

            db.session.commit()
            flash('Cotizacion guardada exitosamente', 'success')
            return redirect(url_for('clientes.formularios'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar: {str(e)}', 'danger')

    return render_template(
        'clientes/form_freight_quote.html',
        form_data=form_data
    )


@clientes_bp.route('/formularios/freight-quote/<int:id>/descargar')
@login_required
def descargar_freight_quote(id):
    """Descargar Excel con datos del Freight Quote"""
    form = FreightQuoteForm.query.get_or_404(id)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        import os

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotizacion"

        # Configurar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12

        # Estilos
        header_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        currency_font = Font(size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Logo (si existe)
        logo_path = r'c:\Users\nicol\OneDrive - The Coca-Cola Company\Proyectos\Lux\lux_portal\static\img\freightwise_logo.png'
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 200
                img.height = 80
                ws.add_image(img, 'C1')
            except:
                pass

        # Titulo de ruta
        ws.merge_cells('A7:I7')
        ws['A7'] = form.ruta or 'RUTA'
        ws['A7'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A7'].fill = header_fill
        ws['A7'].alignment = Alignment(horizontal='center')

        # Headers de la tabla de costos
        headers = ['NET WEIGHT', 'VOLUME WEIGHT', 'RATE', 'WeightxRate', 'DUE AGENT', 'DUE CARRIER', 'CERTIFICATE', 'PHYTO', 'TOTAL']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Datos de costos
        values = [
            form.net_weight,
            form.volume_weight,
            form.rate,
            form.weight_x_rate,
            form.due_agent,
            form.due_carrier,
            form.certificate,
            form.phyto,
            form.total
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=9, column=col)
            if col == 1:  # Net weight (sin $)
                cell.value = value
            elif col in [2, 4, 5, 6, 7, 8, 9]:  # Con $
                cell.value = f"$ {value:,.2f}"
            else:  # Rate
                cell.value = f"$ {value:,.2f}"
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # NOTES section
        ws.merge_cells('A11:I11')
        ws['A11'] = 'NOTES'
        ws['A11'].font = Font(bold=True, color="FFFFFF")
        ws['A11'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        ws['A11'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A12:I14')
        ws['A12'] = form.notes or ''
        ws['A12'].alignment = Alignment(wrap_text=True, vertical='top')

        # ACCOUNT section
        ws.merge_cells('A16:I16')
        ws['A16'] = 'ACCOUNT'
        ws['A16'].font = Font(bold=True, color="FFFFFF")
        ws['A16'].fill = header_fill
        ws['A16'].alignment = Alignment(horizontal='center')

        # Datos bancarios
        bank_data = [
            ('BANK', form.bank_name),
            ('SWIFT', form.bank_swift),
            ('ACCOUNT', form.bank_account),
            ('COMPANY', form.company_name),
            ('TAXID', form.tax_id),
            ('ADDRESS', form.company_address),
        ]

        for i, (label, value) in enumerate(bank_data, 17):
            ws.cell(row=i, column=1).value = label
            ws.cell(row=i, column=1).font = Font(bold=True)
            ws.merge_cells(f'B{i}:I{i}')
            ws.cell(row=i, column=2).value = value
            ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')

        wb.save(output)
        output.seek(0)

        filename = f'FreightQuote_{form.ruta}_{form.id}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'danger')
        return redirect(url_for('clientes.formularios'))


@clientes_bp.route('/formularios/freight-quote/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_freight_quote(id):
    """Eliminar formulario Freight Quote"""
    try:
        form = FreightQuoteForm.query.get_or_404(id)
        db.session.delete(form)
        db.session.commit()
        flash('Cotizacion eliminada', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('clientes.formularios'))


# ============================================================================
# FUNCIONES AUXILIARES PARA PARSEAR EXCEL
# ============================================================================

def parse_customer_associate_excel(df):
    """Parsea el Excel FW CUSTOMER ASSOCIATE FORMAT"""
    data = {}

    def get_cell_value(row_idx, col_idx=2):
        """Obtiene valor de celda de forma segura"""
        try:
            val = df.iloc[row_idx, col_idx]
            return str(val).strip() if pd.notna(val) else ''
        except:
            return ''

    # General Information
    data['legal_name'] = get_cell_value(9)
    data['tax_id'] = get_cell_value(10)
    data['company_type'] = get_cell_value(11)
    data['address'] = get_cell_value(12)
    data['city_country'] = get_cell_value(13)
    data['telephone'] = get_cell_value(14)
    data['email'] = get_cell_value(15)
    data['website'] = get_cell_value(16)
    data['business_activity'] = get_cell_value(17)
    data['annual_revenue'] = get_cell_value(18)

    # Legal Representative
    data['legal_rep_name'] = get_cell_value(20)
    data['legal_rep_id'] = get_cell_value(21)
    data['legal_rep_telephone'] = get_cell_value(22)
    data['legal_rep_email'] = get_cell_value(23)

    # Shareholders
    shareholders = []
    for i in range(28, 31):
        name = get_cell_value(i, 1)
        sid = get_cell_value(i, 2)
        if name and name != 'NAME':
            shareholders.append({'name': name, 'id_number': sid})
    data['shareholders'] = shareholders

    # Financial
    data['financial_contact'] = get_cell_value(32)
    data['bank_name'] = get_cell_value(33)
    data['bank_account'] = get_cell_value(34)
    data['bank_address'] = get_cell_value(35)
    data['swift_aba'] = get_cell_value(36)

    return data


def parse_shipping_instructions_excel(df):
    """Parsea el Excel FWCustomer Shipping Instructions"""
    data = {}

    def get_cell_value(row_idx, col_idx=1):
        """Obtiene valor de celda de forma segura"""
        try:
            val = df.iloc[row_idx, col_idx]
            return str(val).strip() if pd.notna(val) else ''
        except:
            return ''

    # Contact Information
    data['contact_full_name'] = get_cell_value(8)
    data['contact_job_title'] = get_cell_value(9)
    data['contact_email'] = get_cell_value(10)
    data['contact_phone'] = get_cell_value(11)
    data['contact_whatsapp'] = get_cell_value(12)

    # Billing
    data['billing_address'] = get_cell_value(15)

    # Collect multiple billing emails
    billing_emails = []
    for i in range(16, 21):
        email = get_cell_value(i)
        if email and '@' in email:
            billing_emails.append(email)
    data['billing_emails'] = '\n'.join(billing_emails)

    # Additional
    data['uses_customs_broker'] = get_cell_value(23)
    data['customs_broker_info'] = get_cell_value(24)
    data['customs_broker_emails'] = get_cell_value(25)

    # Documents
    documents = {}
    for i in range(29, 35):
        label = get_cell_value(i, 0).lower()
        value = get_cell_value(i, 1)
        if 'invoice' in label and 'master' not in label:
            documents['invoice'] = bool(value and value.lower() == 'x')
        elif 'coa' in label:
            documents['coa'] = bool(value and value.lower() == 'x')
        elif 'phyto' in label and 'destination' not in label:
            documents['phyto'] = bool(value and value.lower() == 'x')
        elif 'phyto destination' in label:
            documents['phyto_destination'] = bool(value and value.lower() == 'x')
        elif 'master' in label:
            documents['master_invoice'] = bool(value and value.lower() == 'x')
    data['documents'] = documents

    return data


def parse_freight_quote_excel(df):
    """Parsea el Excel de Freight Quote de FreightWise"""
    data = {}

    def get_cell_value(row_idx, col_idx=0):
        """Obtiene valor de celda de forma segura"""
        try:
            val = df.iloc[row_idx, col_idx]
            if pd.notna(val):
                # Limpiar el valor de simbolos de moneda
                val_str = str(val).strip()
                val_str = val_str.replace('$', '').replace(',', '').strip()
                return val_str
            return ''
        except:
            return ''

    def get_float(row_idx, col_idx=0):
        """Obtiene valor numerico de celda"""
        try:
            val = get_cell_value(row_idx, col_idx)
            if val:
                return float(val)
            return 0.0
        except:
            return 0.0

    # Buscar la ruta (primera celda que contenga un patron como UIO-BKK)
    for row in range(20):
        for col in range(10):
            val = get_cell_value(row, col)
            if val and '-' in val and len(val) <= 10 and val.isupper():
                data['ruta'] = val
                break
        if 'ruta' in data:
            break

    # Buscar los valores de costos basandose en headers conocidos
    # Buscamos en las filas cercanas al header
    for row in range(30):
        row_values = [get_cell_value(row, col) for col in range(10)]
        row_text = ' '.join(row_values).upper()

        # Si encontramos NET WEIGHT en esta fila, los datos estan en la siguiente
        if 'NET WEIGHT' in row_text or 'NETWEIGHT' in row_text:
            data['net_weight'] = get_float(row + 1, 0)
            data['volume_weight'] = get_float(row + 1, 1)
            data['rate'] = get_float(row + 1, 2)
            data['weight_x_rate'] = get_float(row + 1, 3)
            data['due_agent'] = get_float(row + 1, 4)
            data['due_carrier'] = get_float(row + 1, 5)
            data['certificate'] = get_float(row + 1, 6)
            data['phyto'] = get_float(row + 1, 7)
            data['total'] = get_float(row + 1, 8)
            break

    # Buscar NOTES
    for row in range(30):
        val = get_cell_value(row, 0)
        if val and 'NOTE' in val.upper():
            # El contenido de las notas esta en las filas siguientes
            notes_parts = []
            for i in range(1, 5):
                note_val = get_cell_value(row + i, 0)
                if note_val and 'ACCOUNT' not in note_val.upper():
                    notes_parts.append(note_val)
            data['notes'] = ' '.join(notes_parts)
            break

    # Buscar datos bancarios
    for row in range(40):
        val = get_cell_value(row, 0)
        if val:
            val_upper = val.upper()
            if val_upper == 'BANK':
                data['bank_name'] = get_cell_value(row, 1) or get_cell_value(row, 2)
            elif val_upper == 'SWIFT':
                data['bank_swift'] = get_cell_value(row, 1) or get_cell_value(row, 2)
            elif val_upper == 'ACCOUNT':
                data['bank_account'] = get_cell_value(row, 1) or get_cell_value(row, 2)
            elif val_upper == 'COMPANY':
                data['company_name'] = get_cell_value(row, 1) or get_cell_value(row, 2)
            elif val_upper == 'TAXID':
                data['tax_id'] = get_cell_value(row, 1) or get_cell_value(row, 2)
            elif val_upper == 'ADDRESS':
                data['company_address'] = get_cell_value(row, 1) or get_cell_value(row, 2)

    return data


# ============================================================================
# PDF PARSERS
# ============================================================================

def _extract_pdf_text(file_obj):
    """Extrae todo el texto de un PDF usando pdfplumber"""
    import pdfplumber
    text = ''
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + '\n'
    return text


def _find_value_after(text, label, default=''):
    """Busca un label en el texto y retorna el valor que sigue"""
    import re
    # Try: "Label: value" or "Label value"
    pattern = re.compile(re.escape(label) + r'\s*[:\-]?\s*(.+)', re.IGNORECASE)
    match = pattern.search(text)
    if match:
        val = match.group(1).strip()
        # Stop at next label (uppercase word followed by colon)
        val = re.split(r'\n', val)[0].strip()
        return val
    return default


def parse_customer_associate_pdf(file_obj):
    """Parsea un PDF de Customer Associate"""
    text = _extract_pdf_text(file_obj)
    data = {}

    # Map common labels to fields
    label_map = {
        'legal name': 'legal_name',
        'company name': 'legal_name',
        'tax id': 'tax_id',
        'ruc': 'tax_id',
        'company type': 'company_type',
        'address': 'address',
        'city': 'city_country',
        'telephone': 'telephone',
        'phone': 'telephone',
        'email': 'email',
        'website': 'website',
        'business activity': 'business_activity',
        'annual revenue': 'annual_revenue',
        'legal representative': 'legal_rep_name',
        'representative name': 'legal_rep_name',
        'representative id': 'legal_rep_id',
        'representative telephone': 'legal_rep_telephone',
        'representative email': 'legal_rep_email',
        'bank name': 'bank_name',
        'bank': 'bank_name',
        'account': 'bank_account',
        'bank account': 'bank_account',
        'bank address': 'bank_address',
        'swift': 'swift_aba',
        'aba': 'swift_aba',
    }

    for label, field in label_map.items():
        val = _find_value_after(text, label)
        if val and field not in data:
            data[field] = val

    return data


def parse_shipping_pdf(file_obj):
    """Parsea un PDF de Shipping Instructions"""
    text = _extract_pdf_text(file_obj)
    data = {}

    label_map = {
        'full name': 'contact_full_name',
        'contact name': 'contact_full_name',
        'job title': 'contact_job_title',
        'email': 'contact_email',
        'phone': 'contact_phone',
        'telephone': 'contact_phone',
        'whatsapp': 'contact_whatsapp',
        'billing address': 'billing_address',
        'customs broker': 'uses_customs_broker',
    }

    for label, field in label_map.items():
        val = _find_value_after(text, label)
        if val and field not in data:
            data[field] = val

    return data
