import io
import os
import json
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file
from lux_portal.current_status import current_status_bp
from lux_portal.current_status.models import StatusClient, AirlineRate, StatusAirline, StatusPayment, ClientShipment, ShipmentHistory
from lux_portal.extensions import db
from lux_portal.auth.decorators import login_required


def get_logo_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(current_dir, '..', 'static', 'images', 'freightwise_logo.png')


PAYMENT_FIELDS = ['route', 'net_weight', 'volume_weight', 'rate',
                  'pay_due_agent', 'pay_due_carrier', 'certificate', 'pay_phyto', 'pay_notes']


_payment_cols_checked = False


def _ensure_payment_columns():
    """Add new payment columns if they don't exist yet (no-migration approach)."""
    global _payment_cols_checked
    if _payment_cols_checked:
        return
    _payment_cols_checked = True
    new_cols = {
        'route': "VARCHAR(200) DEFAULT ''",
        'net_weight': "VARCHAR(100) DEFAULT ''",
        'volume_weight': "VARCHAR(100) DEFAULT ''",
        'rate': "VARCHAR(100) DEFAULT ''",
        'pay_due_agent': "VARCHAR(100) DEFAULT '50'",
        'pay_due_carrier': "VARCHAR(100) DEFAULT '25'",
        'certificate': "VARCHAR(100) DEFAULT '15'",
        'pay_phyto': "VARCHAR(100) DEFAULT '2.50'",
        'pay_notes': "TEXT DEFAULT ''",
    }
    from sqlalchemy import text, inspect
    try:
        inspector = inspect(db.engine)
        existing = {c['name'] for c in inspector.get_columns('status_payments')}
        for col_name, col_type in new_cols.items():
            if col_name not in existing:
                db.session.execute(text(
                    f'ALTER TABLE status_payments ADD COLUMN {col_name} {col_type}'
                ))
        db.session.commit()
    except Exception:
        db.session.rollback()


# ===================== PAGES =====================

@current_status_bp.route('/')
@login_required
def dashboard():
    """Dashboard con lista de clientes y resumen de pagos"""
    busqueda = request.args.get('q', '').strip()
    query = StatusClient.query.filter_by(activo=True)
    if busqueda:
        query = query.filter(StatusClient.nombre.ilike(f'%{busqueda}%'))
    clientes = query.order_by(StatusClient.fecha_actualizacion.desc()).all()
    return render_template('current_status/dashboard.html', clientes=clientes, busqueda=busqueda)


@current_status_bp.route('/facturacion')
@login_required
def facturacion():
    """Vista de facturacion con tablas Shipment/Facturacion/Costos de todos los clientes"""
    busqueda = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 100

    # Build client filter
    client_query = StatusClient.query.filter_by(activo=True)
    if busqueda:
        client_query = client_query.filter(StatusClient.nombre.ilike(f'%{busqueda}%'))
    client_ids = [c.id for c in client_query.all()]

    # Paginated shipments across all matching clients
    shipments_q = ClientShipment.query.filter(ClientShipment.client_id.in_(client_ids)) if client_ids else ClientShipment.query.filter(False)
    total = shipments_q.count()
    shipments = shipments_q.order_by(ClientShipment.id).offset((page - 1) * per_page).limit(per_page).all()
    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template('current_status/facturacion.html',
                           shipments=shipments, busqueda=busqueda,
                           page=page, total_pages=total_pages, total=total, per_page=per_page)


@current_status_bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_cliente():
    """Crear nuevo cliente"""
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        if not nombre:
            flash('El nombre del cliente es requerido', 'danger')
            return redirect(url_for('current_status.nuevo_cliente'))
        cliente = StatusClient(nombre=nombre)
        db.session.add(cliente)
        db.session.commit()
        flash(f'Cliente {nombre} creado exitosamente', 'success')
        return redirect(url_for('current_status.detalle_cliente', id=cliente.id))
    return render_template('current_status/nuevo.html')


@current_status_bp.route('/<int:id>')
@login_required
def detalle_cliente(id):
    """Ver detalle del cliente con tablas editables"""
    _ensure_payment_columns()
    cliente = StatusClient.query.get_or_404(id)
    _migrate_legacy_addcosts(cliente)
    # Only show the last shipment in detail; all others are in historial
    last = ClientShipment.query.filter_by(client_id=id).order_by(ClientShipment.id.desc()).first()
    display_shipments = [last] if last else []
    total_shipments = ClientShipment.query.filter_by(client_id=id).count()
    return render_template('current_status/detail.html', cliente=cliente,
                           display_shipments=display_shipments, total_shipments=total_shipments)


def _migrate_legacy_addcosts(cliente):
    """Convert legacy additional_costs columns to addcosts array in extra_data."""
    import re
    dirty = False
    for rate in cliente.rates:
        extra = rate.get_extra()
        if extra.get('addcosts'):
            continue  # already migrated
        names_str = (rate.additional_costs or '').strip()
        values_str = (rate.additional_costs_value or '').strip()
        if not names_str and not values_str:
            continue
        # Parse values: split by $ sign to get individual amounts
        values = re.findall(r'\$?([\d.,]+)', values_str)
        n_values = len(values) if values else 1
        # Parse names: split by spaces, last (n-1) tokens are individual names,
        # remaining tokens joined form the first name
        name_tokens = names_str.split() if names_str else []
        if n_values <= 1:
            names = [names_str]
        elif len(name_tokens) <= n_values:
            names = name_tokens
        else:
            tail = name_tokens[-(n_values - 1):]
            head = ' '.join(name_tokens[:len(name_tokens) - (n_values - 1)])
            names = [head] + tail
        # Build addcosts array
        addcosts = []
        for i in range(max(len(names), len(values))):
            name = names[i] if i < len(names) else ''
            val = values[i] if i < len(values) else ''
            addcosts.append({'name': name, 'value': val})
        extra['addcosts'] = addcosts
        rate.extra_data = json.dumps(extra)
        rate.additional_costs = ''
        rate.additional_costs_value = ''
        dirty = True
    if dirty:
        db.session.commit()


# ===================== API: CLIENT =====================

@current_status_bp.route('/api/client/<int:id>/estado', methods=['PUT'])
@login_required
def cambiar_estado(id):
    cliente = StatusClient.query.get_or_404(id)
    data = request.get_json()
    cliente.estado = data.get('estado', 'pendiente')
    db.session.commit()
    return jsonify({'success': True, 'estado': cliente.estado})


@current_status_bp.route('/api/client/<int:id>/nombre', methods=['PUT'])
@login_required
def cambiar_nombre(id):
    cliente = StatusClient.query.get_or_404(id)
    data = request.get_json()
    cliente.nombre = data.get('nombre', cliente.nombre)
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/client/<int:id>', methods=['DELETE'])
@login_required
def eliminar_cliente(id):
    cliente = StatusClient.query.get_or_404(id)
    cliente.activo = False
    db.session.commit()
    return jsonify({'success': True})


# ===================== API: AIRLINE RATES (Table 1) =====================

@current_status_bp.route('/api/client/<int:id>/rate', methods=['POST'])
@login_required
def agregar_rate(id):
    StatusClient.query.get_or_404(id)
    rate = AirlineRate(client_id=id)
    db.session.add(rate)
    db.session.commit()
    return jsonify({'success': True, 'id': rate.id})


@current_status_bp.route('/api/rate/<int:id>', methods=['PUT'])
@login_required
def actualizar_rate(id):
    rate = AirlineRate.query.get_or_404(id)
    data = request.get_json()
    if 'extra' in data:
        extra = rate.get_extra()
        extra.update(data['extra'])
        rate.extra_data = json.dumps(extra)
    else:
        for field in ['airline', 'route', 'transit_time', 'kg_availability', 'date',
                      'net_rate', 'operative', 'net_ops', 'profit',
                      'final_rate', 'additional_costs', 'additional_costs_value', 'notes']:
            if field in data:
                setattr(rate, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/rate/<int:id>', methods=['DELETE'])
@login_required
def eliminar_rate(id):
    rate = AirlineRate.query.get_or_404(id)
    db.session.delete(rate)
    db.session.commit()
    return jsonify({'success': True})


# ===================== API: STATUS AIRLINES (Table 2) =====================

@current_status_bp.route('/api/client/<int:id>/airline', methods=['POST'])
@login_required
def agregar_airline(id):
    StatusClient.query.get_or_404(id)
    airline = StatusAirline(client_id=id)
    db.session.add(airline)
    db.session.commit()
    return jsonify({'success': True, 'id': airline.id})


@current_status_bp.route('/api/airline/<int:id>', methods=['PUT'])
@login_required
def actualizar_airline(id):
    airline = StatusAirline.query.get_or_404(id)
    data = request.get_json()
    if 'extra' in data:
        extra = airline.get_extra()
        extra.update(data['extra'])
        airline.extra_data = json.dumps(extra)
    else:
        for field in ['current_status', 'proximo_vuelo', 'kg', 'entrega_fincas', 'hora_maxima', 'aerolinea', 'all_in_rate']:
            if field in data:
                setattr(airline, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/airline/<int:id>', methods=['DELETE'])
@login_required
def eliminar_airline(id):
    airline = StatusAirline.query.get_or_404(id)
    db.session.delete(airline)
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/client/<int:id>/upload-airlines', methods=['POST'])
@login_required
def upload_airlines_excel(id):
    """Subir Excel para llenar la tabla Current Status (airlines)."""
    import pandas as pd

    client = StatusClient.query.get_or_404(id)

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se envio archivo'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': 'Archivo vacio'}), 400

    try:
        df = pd.read_excel(file, header=0)
        df.columns = [str(c).strip() for c in df.columns]

        # Mapeo flexible de columnas del Excel a campos del modelo
        col_map = {
            'current_status': ['current status', 'current_status', 'status', 'estado'],
            'proximo_vuelo': ['next flight', 'proximo vuelo', 'proximo_vuelo', 'vuelo', 'flight'],
            'kg': ['kg', 'kilos', 'weight', 'peso'],
            'entrega_fincas': ['farms deliver', 'entrega fincas', 'entrega_fincas', 'fincas', 'farms'],
            'hora_maxima': ['farm maximum delivery time', 'hora maxima', 'hora_maxima', 'max delivery', 'hora max'],
            'aerolinea': ['airline', 'aerolinea', 'aerolínea'],
            'all_in_rate': ['all in rate', 'all_in_rate', 'rate', 'tarifa'],
        }

        # Encontrar mapeo real: campo_modelo -> columna_excel
        field_to_excel = {}
        df_lower = {str(c).lower(): c for c in df.columns}
        for field, aliases in col_map.items():
            for alias in aliases:
                if alias.lower() in df_lower:
                    field_to_excel[field] = df_lower[alias.lower()]
                    break

        if not field_to_excel:
            return jsonify({'success': False, 'error': 'No se encontraron columnas reconocibles en el Excel. Columnas esperadas: Current Status, Next Flight, KG, Farms Deliver, Airline, All in Rate'}), 400

        # Columnas extra (las que no se mapearon a campos conocidos)
        mapped_excel_cols = set(field_to_excel.values())
        extra_cols = [c for c in df.columns if c not in mapped_excel_cols]

        created = 0
        for _, row in df.iterrows():
            # Saltar filas completamente vacias
            if all(pd.isna(row.get(field_to_excel.get(f, ''), '')) for f in col_map if f in field_to_excel):
                continue

            airline = StatusAirline(client_id=client.id)
            for field, excel_col in field_to_excel.items():
                val = row.get(excel_col, '')
                setattr(airline, field, str(val).strip() if pd.notna(val) else '')

            # Extra cols -> extra_data JSON
            if extra_cols:
                extra = {}
                for ec in extra_cols:
                    val = row.get(ec, '')
                    if pd.notna(val) and str(val).strip():
                        extra[ec] = str(val).strip()
                if extra:
                    airline.extra_data = json.dumps(extra)

            db.session.add(airline)
            created += 1

        db.session.commit()

        return jsonify({'success': True, 'created': created, 'fields': list(field_to_excel.keys())})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@current_status_bp.route('/api/client/<int:id>/upload-rates', methods=['POST'])
@login_required
def upload_rates_excel(id):
    """Subir Excel de cotizacion para llenar la tabla Airline Rates (Table 1)."""
    import openpyxl

    client = StatusClient.query.get_or_404(id)

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se envio archivo'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': 'Archivo vacio'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Buscar fila de encabezados (contiene "AIRLINE" en columna A)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=20):
            val = str(row[0].value or '').strip().upper()
            if val == 'AIRLINE':
                header_row = row[0].row
                break

        if not header_row:
            return jsonify({'success': False, 'error': 'No se encontro la fila de encabezados (AIRLINE). Verifique el formato del Excel.'}), 400

        # Leer encabezados y mapear indices de columnas
        headers = []
        for cell in ws[header_row]:
            headers.append(str(cell.value or '').strip().upper().replace('\n', ' '))

        # Mapeo de encabezado de cotizacion -> campo AirlineRate
        # Usar lista ordenada de mayor a menor longitud para evitar conflictos
        # (ej: "RATE INCREASE" no debe matchear con "RATE")
        header_mappings = [
            ('TRANSIT TIME', 'transit_time'),
            ('ADD CHARGES', 'additional_costs'),
            ('RATE INCREASE', None),  # Ignorar, no aplica a Table 1
            ('DATE', None),           # Ignorar, no aplica
            ('AIRLINE', 'airline'),
            ('ITINERARY', 'route'),
            ('KG', 'kg_availability'),
            ('RATE', 'final_rate'),
            ('NOTES', 'notes'),
        ]

        # Encontrar indice de cada campo
        col_indices = {}  # field_name -> col_index
        add_charges_idx = None  # columna L (nombre del cargo)
        add_charges_val_idx = None  # columna M (valor del cargo)
        matched_cols = set()
        for i, h in enumerate(headers):
            if i in matched_cols:
                continue
            for header_name, field_name in header_mappings:
                if h == header_name or (header_name in h and i not in matched_cols):
                    matched_cols.add(i)
                    if field_name:
                        col_indices[field_name] = i
                    if header_name == 'ADD CHARGES':
                        add_charges_idx = i
                        add_charges_val_idx = i + 1  # siguiente columna = valor
                    break

        def cell_val(row_cells, idx):
            """Obtener valor de celda como string limpio."""
            if idx is None or idx >= len(row_cells):
                return ''
            v = row_cells[idx].value
            if v is None:
                return ''
            return str(v).strip()

        # Leer filas de datos despues del header
        rates_list = []
        current_rate = None
        last_airline = ''  # Para heredar airline en rutas multiples

        for row in ws.iter_rows(min_row=header_row + 1):
            cells = list(row)
            airline_val = cell_val(cells, col_indices.get('airline', 0))
            itinerary_val = cell_val(cells, col_indices.get('route', 2))
            rate_val = cell_val(cells, col_indices.get('final_rate', 8))

            # Parar si llegamos a "FreightWise Additional Charges"
            if 'freightwise' in airline_val.lower() or 'freight wise' in airline_val.lower() or 'additional charges' in airline_val.lower():
                break

            # Detectar si es fila principal (tiene itinerary o rate)
            is_main_row = bool(airline_val or itinerary_val or rate_val)

            # Detectar si es sub-fila de cargos adicionales
            charge_name = cell_val(cells, add_charges_idx) if add_charges_idx is not None else ''
            charge_val = cell_val(cells, add_charges_val_idx) if add_charges_val_idx is not None else ''

            if is_main_row:
                # Si tiene airline, actualizar last_airline
                if airline_val:
                    last_airline = airline_val
                # Nueva fila principal -> crear nuevo rate
                current_rate = {
                    'airline': airline_val if airline_val else last_airline,
                    'route': itinerary_val,
                    'transit_time': cell_val(cells, col_indices.get('transit_time', 3)),
                    'kg_availability': cell_val(cells, col_indices.get('kg_availability', 7)),
                    'final_rate': rate_val,
                    'notes': cell_val(cells, col_indices.get('notes', 14)),
                    'additional_costs': '',
                }
                # Recoger cargo adicional si existe en la misma fila
                if charge_name:
                    current_rate['additional_costs'] = f"{charge_name}: {charge_val}" if charge_val else charge_name
                rates_list.append(current_rate)

            elif charge_name and current_rate:
                # Sub-fila con cargo adicional -> agregar al rate anterior
                extra = f"{charge_name}: {charge_val}" if charge_val else charge_name
                if current_rate['additional_costs']:
                    current_rate['additional_costs'] += f"\n{extra}"
                else:
                    current_rate['additional_costs'] = extra

        # Guardar en base de datos
        created = 0
        for data in rates_list:
            rate = AirlineRate(client_id=client.id)
            for field, val in data.items():
                if val:
                    setattr(rate, field, val)
            db.session.add(rate)
            created += 1

        db.session.commit()

        return jsonify({'success': True, 'created': created})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ===================== API: PAYMENTS (Table 3) =====================

@current_status_bp.route('/api/client/<int:id>/payment', methods=['POST'])
@login_required
def agregar_payment(id):
    StatusClient.query.get_or_404(id)
    payment = StatusPayment(
        client_id=id,
        pay_due_agent='50',
        pay_due_carrier='25',
        certificate='15',
        pay_phyto='2.50'
    )
    db.session.add(payment)
    db.session.commit()
    return jsonify({'success': True, 'id': payment.id})


@current_status_bp.route('/api/payment/<int:id>', methods=['PUT'])
@login_required
def actualizar_payment(id):
    payment = StatusPayment.query.get_or_404(id)
    data = request.get_json()
    if 'extra' in data:
        extra = payment.get_extra()
        extra.update(data['extra'])
        payment.extra_data = json.dumps(extra)
    else:
        for field in ['valor', 'fecha', 'credito'] + PAYMENT_FIELDS:
            if field in data:
                setattr(payment, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/payment/<int:id>', methods=['DELETE'])
@login_required
def eliminar_payment(id):
    payment = StatusPayment.query.get_or_404(id)
    db.session.delete(payment)
    db.session.commit()
    return jsonify({'success': True})


# ===================== API: SHIPMENTS (Table 4 & 5) =====================

SHIPMENT_FIELDS = ['ship_date', 'transport', 'mark', 'awb', 'client_name', 'origin', 'destination',
                   'airline', 'fulles', 'piezas', 'pieces_gross', 'volume', 'charge', 'phyto',
                   'dup_phyto', 'c_origin', 'dup_co', 'termografo', 'transmision']

FACTURA_FIELDS = ['facturacion', 'handling_juni', 'flete', 'fsc', 'esc', 'due_agent', 'due_carrier',
                  'fito_venta', 'co_venta', 'termografo_venta', 'fitos_venta', 'dup_fito_venta',
                  'co_factura', 'dup_co_factura', 'termografo_factura',
                  'transmision_factura', 'beneficio', 'beneficio_x_kg']

COSTOS_FIELDS = ['costos', 'costo_bodega', 'costo_guia', 'flete_costo', 'due_carrier_costo',
                 'fsc_costo', 'esc_costo', 'costo_x_kg', 'costo_bod_unit', 'costo_guia_unit',
                 'fito_costo_unit', 'co_costo_unit', 'termografo_costo_unit', 'fitos_costo',
                 'dup_fitos_costo', 'termografo_costo', 'co_costo', 'dup_co_costo',
                 'transmision_costo', 'costos_fijos', 'utilidad']


@current_status_bp.route('/api/client/<int:id>/shipment', methods=['POST'])
@login_required
def agregar_shipment(id):
    client = StatusClient.query.get_or_404(id)
    shipment = ClientShipment(
        client_id=id,
        mark='NO',
        client_name=client.nombre.upper() if client.nombre else '',
        phyto='1',
        dup_phyto='0',
        c_origin='1',
        dup_co='0',
        termografo='0',
        transmision='1',
        due_agent='50',
        fito_venta='2.50',
        co_venta='15',
        termografo_venta='35',
        costo_bod_unit='0.069',
        costo_guia_unit='0.005',
        fito_costo_unit='1.7',
        co_costo_unit='13'
    )
    db.session.add(shipment)
    db.session.commit()
    return jsonify({'success': True, 'id': shipment.id})


@current_status_bp.route('/api/shipment/<int:id>', methods=['PUT'])
@login_required
def actualizar_shipment(id):
    shipment = ClientShipment.query.get_or_404(id)
    data = request.get_json()
    for field in SHIPMENT_FIELDS + FACTURA_FIELDS + COSTOS_FIELDS:
        if field in data:
            setattr(shipment, field, data[field])
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/shipment/<int:id>', methods=['DELETE'])
@login_required
def eliminar_shipment(id):
    shipment = ClientShipment.query.get_or_404(id)
    db.session.delete(shipment)
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/client/<int:id>/clear-shipments', methods=['DELETE'])
@login_required
def clear_shipments(id):
    """Delete ALL shipments for a client at once."""
    StatusClient.query.get_or_404(id)
    count = ClientShipment.query.filter_by(client_id=id).delete()
    db.session.commit()
    return jsonify({'success': True, 'deleted': count})


@current_status_bp.route('/api/client/<int:id>/enviar-shipments', methods=['POST'])
@login_required
def enviar_shipments(id):
    """Enviar embarques al historial y contar"""
    cliente = StatusClient.query.get_or_404(id)
    shipments = ClientShipment.query.filter_by(client_id=id).all()
    if not shipments:
        return jsonify({'success': False, 'error': 'No hay embarques para enviar'}), 400

    # Snapshot data
    snapshot = []
    for s in shipments:
        row = {}
        for f in SHIPMENT_FIELDS + FACTURA_FIELDS + COSTOS_FIELDS:
            row[f] = getattr(s, f, '') or ''
        snapshot.append(row)

    history = ShipmentHistory(
        client_id=id,
        client_name=cliente.nombre,
        shipment_count=len(shipments),
        shipment_data=json.dumps(snapshot)
    )
    db.session.add(history)
    db.session.commit()

    return jsonify({
        'success': True,
        'history_id': history.id,
        'count': len(shipments)
    })


@current_status_bp.route('/historial')
@login_required
def historial_clientes():
    """Historial: todos los embarques de todos los clientes con paginacion"""
    busqueda = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 100

    # Build client filter
    client_query = StatusClient.query.filter_by(activo=True)
    if busqueda:
        client_query = client_query.filter(StatusClient.nombre.ilike(f'%{busqueda}%'))
    client_ids = [c.id for c in client_query.all()]

    # Paginated shipments across all matching clients
    shipments_q = ClientShipment.query.filter(ClientShipment.client_id.in_(client_ids)) if client_ids else ClientShipment.query.filter(False)
    total = shipments_q.count()
    shipments = shipments_q.order_by(ClientShipment.id).offset((page - 1) * per_page).limit(per_page).all()
    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template(
        'current_status/historial.html',
        shipments=shipments, busqueda=busqueda,
        page=page, total_pages=total_pages, total=total, per_page=per_page
    )


@current_status_bp.route('/api/historial/<int:id>', methods=['DELETE'])
@login_required
def eliminar_historial(id):
    """Eliminar una entrada completa del historial"""
    entry = ShipmentHistory.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    return jsonify({'success': True})


@current_status_bp.route('/api/historial/<int:id>/row/<int:row_idx>', methods=['DELETE'])
@login_required
def eliminar_historial_row(id, row_idx):
    """Eliminar una fila individual de una entrada del historial"""
    entry = ShipmentHistory.query.get_or_404(id)
    shipments = entry.get_shipments()
    if row_idx < 0 or row_idx >= len(shipments):
        return jsonify({'success': False, 'error': 'Fila no encontrada'}), 404
    shipments.pop(row_idx)
    if not shipments:
        db.session.delete(entry)
    else:
        entry.shipment_data = json.dumps(shipments)
        entry.shipment_count = len(shipments)
    db.session.commit()
    return jsonify({'success': True})


# ===================== API: CUSTOM COLUMNS =====================

@current_status_bp.route('/api/client/<int:id>/custom-column', methods=['POST'])
@login_required
def agregar_custom_column(id):
    cliente = StatusClient.query.get_or_404(id)
    data = request.get_json()
    table_type = data.get('table_type')
    col_name = data.get('name', 'Nueva Columna')
    try:
        cols = json.loads(cliente.custom_columns or '{}')
        if table_type not in cols:
            cols[table_type] = []
        cols[table_type].append(col_name)
        cliente.custom_columns = json.dumps(cols)
        db.session.commit()
        return jsonify({'success': True, 'columns': cols[table_type], 'index': len(cols[table_type]) - 1})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@current_status_bp.route('/api/client/<int:id>/custom-column', methods=['DELETE'])
@login_required
def eliminar_custom_column(id):
    cliente = StatusClient.query.get_or_404(id)
    data = request.get_json()
    table_type = data.get('table_type')
    col_name = data.get('name')
    try:
        cols = json.loads(cliente.custom_columns or '{}')
        if table_type in cols and col_name in cols[table_type]:
            cols[table_type].remove(col_name)
            cliente.custom_columns = json.dumps(cols)
            # Remove data from rows
            if table_type == 'rates':
                rows = cliente.rates
            elif table_type == 'airlines':
                rows = cliente.airlines
            else:
                rows = cliente.payments
            for row in rows:
                extra = row.get_extra()
                extra.pop(col_name, None)
                row.extra_data = json.dumps(extra)
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Column not found'}), 404
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@current_status_bp.route('/api/client/<int:id>/custom-column/rename', methods=['PUT'])
@login_required
def renombrar_custom_column(id):
    cliente = StatusClient.query.get_or_404(id)
    data = request.get_json()
    table_type = data.get('table_type')
    old_name = data.get('old_name')
    new_name = data.get('new_name', '')
    try:
        cols = json.loads(cliente.custom_columns or '{}')
        if table_type in cols and old_name in cols[table_type]:
            idx = cols[table_type].index(old_name)
            cols[table_type][idx] = new_name
            cliente.custom_columns = json.dumps(cols)
            if table_type == 'rates':
                rows = cliente.rates
            elif table_type == 'airlines':
                rows = cliente.airlines
            else:
                rows = cliente.payments
            for row in rows:
                extra = row.get_extra()
                if old_name in extra:
                    extra[new_name] = extra.pop(old_name)
                    row.extra_data = json.dumps(extra)
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Column not found'}), 404
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ===================== EXPORT: SINGLE CLIENT =====================

@current_status_bp.route('/descargar/<int:id>')
@login_required
def descargar_cliente(id):
    """Descargar formulario de un cliente en Excel o PDF"""
    _ensure_payment_columns()
    cliente = StatusClient.query.get_or_404(id)
    formato = request.args.get('formato', 'excel')
    hide_internal = request.args.get('hide_internal', '0') == '1'
    tabla = request.args.get('table', 'all')  # all, rates, status, payment
    if formato == 'pdf':
        return _generar_pdf_cliente(cliente, hide_internal=hide_internal, tabla=tabla)
    return _generar_excel_cliente(cliente, hide_internal=hide_internal, tabla=tabla)


@current_status_bp.route('/descargar-todos')
@login_required
def descargar_todos():
    """Descargar listado de todos los clientes con pagos y estado"""
    clientes = StatusClient.query.filter_by(activo=True).order_by(StatusClient.nombre).all()
    formato = request.args.get('formato', 'excel')
    if formato == 'pdf':
        return _generar_pdf_todos(clientes)
    return _generar_excel_todos(clientes)


# ===================== UPLOAD EXCEL =====================

@current_status_bp.route('/api/client/<int:id>/upload-excel', methods=['POST'])
@login_required
def upload_excel(id):
    """Subir un Excel para rellenar las 3 tablas del cliente"""
    from openpyxl import load_workbook

    cliente = StatusClient.query.get_or_404(id)
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Archivo Excel requerido (.xlsx)'}), 400

    # Fields that should be rounded to 2 decimals
    NUMERIC_FIELDS = {'net_rate', 'operative', 'net_ops', 'profit', 'final_rate',
                      'additional_costs_value', 'kg_availability', 'valor'}

    def round_val(val, field):
        """Round to 2 decimals if field is numeric."""
        if field not in NUMERIC_FIELDS or not val:
            return val
        try:
            return f"{float(val.replace(',', '')):.2f}"
        except (ValueError, AttributeError):
            return val

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Read all rows as lists of string values
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell).strip() if cell is not None else '' for cell in row])

        # Detect table sections by looking for header keywords
        table1_start = None  # Airline Rates
        table2_start = None  # Current Status
        table3_start = None  # Payment

        for i, row in enumerate(rows):
            row_lower = [c.lower() for c in row]
            row_joined = ' '.join(row_lower)

            if table1_start is None and ('airline' in row_lower or 'airline' in row_joined) and ('route' in row_lower or 'route' in row_joined):
                table1_start = i
            elif table1_start is not None and table2_start is None and ('current status' in row_joined or 'current_status' in row_joined or ('proximo' in row_joined and 'vuelo' in row_joined)):
                table2_start = i
            elif table2_start is not None and table3_start is None and ('payment' in row_joined or ('valor' in row_lower and 'fecha' in row_lower)):
                table3_start = i

        # Parse Table 1: Airline Rates (from table1_start+1 to table2_start-1)
        if table1_start is not None:
            end1 = table2_start if table2_start else len(rows)
            for i in range(table1_start + 1, end1):
                row = rows[i]
                # Skip empty rows
                if all(c == '' or c == 'None' for c in row):
                    continue
                rate = AirlineRate(client_id=id)
                # Map columns by position (matching header order)
                fields1 = ['airline', 'route', 'transit_time', 'kg_availability', 'date',
                           'net_rate', 'operative', 'net_ops', 'profit', 'final_rate',
                           'additional_costs', 'additional_costs_value', 'notes']
                for j, field in enumerate(fields1):
                    if j < len(row):
                        val = row[j] if row[j] != 'None' else ''
                        setattr(rate, field, round_val(val, field))
                db.session.add(rate)

        # Parse Table 2: Current Status (from table2_start+1 to table3_start-1)
        if table2_start is not None:
            end2 = table3_start if table3_start else len(rows)
            for i in range(table2_start + 1, end2):
                row = rows[i]
                if all(c == '' or c == 'None' for c in row):
                    continue
                airline = StatusAirline(client_id=id)
                fields2 = ['current_status', 'proximo_vuelo', 'kg', 'entrega_fincas', 'hora_maxima', 'aerolinea', 'all_in_rate']
                for j, field in enumerate(fields2):
                    if j < len(row):
                        val = row[j] if row[j] != 'None' else ''
                        setattr(airline, field, val)
                db.session.add(airline)

        # Parse Table 3: Payment (from table3_start+1 to end)
        if table3_start is not None:
            # Payment header may have: "Payment ClientName", Valor, Fecha, Credito
            # Data rows have: empty first col, valor, fecha, credito
            for i in range(table3_start + 1, len(rows)):
                row = rows[i]
                if all(c == '' or c == 'None' for c in row):
                    continue
                payment = StatusPayment(client_id=id)
                # The payment table header has 4 cols, data starts at col 2 (index 1)
                fields3 = [('valor', 1), ('fecha', 2), ('credito', 3)]
                for field, idx in fields3:
                    if idx < len(row):
                        val = row[idx] if row[idx] != 'None' else ''
                        setattr(payment, field, round_val(val, field))
                db.session.add(payment)

        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ===================== UPLOAD FACTURACION (BULK) =====================

@current_status_bp.route('/api/upload-facturacion', methods=['POST'])
@login_required
def upload_facturacion():
    """Upload FACTURACION Excel and assign shipment data to clients by CUSTOMER match."""
    import tempfile
    from openpyxl import load_workbook
    from collections import defaultdict
    from datetime import date as date_type

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Archivo Excel requerido (.xlsx)'}), 400

    # Save to temp file for reliable reading
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    try:
        file.save(tmp_path)
        tmp.close()

        wb = load_workbook(tmp_path, data_only=True, read_only=True)
        ws = wb['TOTAL OPERACION'] if 'TOTAL OPERACION' in wb.sheetnames else wb.active

        # --- Parse headers (row 1) ---
        headers = None
        customer_col = None
        col_map = {}
        valid_fields = set(SHIPMENT_FIELDS + FACTURA_FIELDS + COSTOS_FIELDS)

        def _cell_to_str(val):
            if val is None:
                return ''
            if isinstance(val, (datetime, date_type)):
                return val.strftime('%d/%m/%Y')
            if isinstance(val, float):
                return str(int(val)) if val == int(val) else f'{val:.2f}'
            if isinstance(val, int):
                return str(val)
            s = str(val).strip()
            return '' if s == 'None' else s

        customers_data = defaultdict(list)

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # Parse headers
                headers = [str(c or '').strip().upper() for c in row]
                # Find CUSTOMER column
                for i, h in enumerate(headers):
                    if 'CUSTOMER' in h or 'CLIENTE' in h:
                        customer_col = i
                        break
                if customer_col is None:
                    wb.close()
                    return jsonify({'success': False, 'error': 'No se encontró columna CUSTOMER/CLIENTE'}), 400

                # Find section boundaries
                factura_col = costos_col = beneficio_col = None
                for i, h in enumerate(headers):
                    if factura_col is None and ('FACTURACION' in h or 'FACTURACIÓN' in h):
                        factura_col = i
                    elif factura_col is not None and costos_col is None and h == 'COSTOS':
                        costos_col = i
                    elif costos_col is not None and beneficio_col is None and h == 'BENEFICIO':
                        beneficio_col = i

                # --- Shipment section (before FACTURACION) ---
                ship_map = {
                    'FECHA': 'ship_date', 'TRANSPORT': 'transport', 'MARK': 'mark',
                    'AWB': 'awb', 'ORIGIN': 'origin', 'DEST': 'destination',
                    'AIRLINE': 'airline', 'FULL': 'fulles', 'PIECES': 'piezas',
                    'GROSS': 'pieces_gross', 'VOL': 'volume', 'CHARGE': 'charge',
                }
                end_ship = factura_col or len(headers)
                for i in range(end_ship):
                    h = headers[i]
                    if h in ship_map:
                        col_map[i] = ship_map[h]
                    elif 'DUP' in h and 'FITO' in h:
                        col_map[i] = 'dup_phyto'
                    elif 'FITO' in h and 'phyto' not in col_map.values():
                        col_map[i] = 'phyto'
                    elif 'DUP' in h and ('C.' in h or 'ORIGEN' in h):
                        col_map[i] = 'dup_co'
                    elif ('C. ORIGEN' in h or 'C.ORIGEN' in h) and 'c_origin' not in col_map.values():
                        col_map[i] = 'c_origin'
                    elif ('TERMOGRAFO' in h or 'TERMÓGRAFO' in h) and 'termografo' not in col_map.values():
                        col_map[i] = 'termografo'
                    elif 'TRANSMISI' in h and 'transmision' not in col_map.values():
                        col_map[i] = 'transmision'

                # --- Facturacion section (positional) ---
                if factura_col is not None:
                    end_fact = costos_col or len(headers)
                    fact_fields = ['facturacion', 'handling_juni', 'flete', 'fsc', 'esc',
                                   'due_agent', 'due_carrier', 'fito_venta', 'co_venta',
                                   'termografo_venta', 'fitos_venta', 'dup_fito_venta',
                                   'co_factura', 'dup_co_factura', 'termografo_factura',
                                   'transmision_factura']
                    for j, i in enumerate(range(factura_col, end_fact)):
                        if j < len(fact_fields):
                            col_map[i] = fact_fields[j]

                # --- Costos section (positional) ---
                if costos_col is not None:
                    end_cost = beneficio_col or len(headers)
                    cost_fields = ['costos', 'costo_bodega', 'costo_guia', 'flete_costo',
                                   'due_carrier_costo', 'fsc_costo', 'esc_costo', 'costo_x_kg',
                                   'costo_bod_unit', 'costo_guia_unit', 'fito_costo_unit',
                                   'co_costo_unit', 'termografo_costo_unit', 'fitos_costo',
                                   'dup_fitos_costo', 'termografo_costo', 'co_costo', 'dup_co_costo',
                                   'transmision_costo']
                    for j, i in enumerate(range(costos_col, end_cost)):
                        if j < len(cost_fields):
                            col_map[i] = cost_fields[j]

                # --- Results section (from BENEFICIO) ---
                if beneficio_col is not None:
                    res_fields = ['beneficio', 'beneficio_x_kg', 'costos_fijos', 'utilidad']
                    for j, i in enumerate(range(beneficio_col, min(beneficio_col + len(res_fields), len(headers)))):
                        if j < len(res_fields):
                            col_map[i] = res_fields[j]
                continue

            # --- Process data rows ---
            cells = list(row)
            if customer_col >= len(cells):
                continue
            customer = _cell_to_str(cells[customer_col])
            if not customer:
                continue

            row_data = {'client_name': customer}
            for ci, field in col_map.items():
                if field not in valid_fields:
                    continue
                if ci < len(cells):
                    row_data[field] = _cell_to_str(cells[ci])
            customers_data[customer.strip().upper()].append(row_data)

        wb.close()

        # --- Find or create clients, add shipments ---
        created = []
        updated = []
        total_shipments = 0

        for customer_key, rows_data in customers_data.items():
            client = StatusClient.query.filter(
                db.func.upper(StatusClient.nombre) == customer_key,
                StatusClient.activo == True
            ).first()

            if client:
                updated.append(client.nombre)
            else:
                client = StatusClient(nombre=rows_data[0]['client_name'])
                db.session.add(client)
                db.session.flush()
                created.append(rows_data[0]['client_name'])

            for row_data in rows_data:
                shipment = ClientShipment(client_id=client.id)
                for field, val in row_data.items():
                    if hasattr(shipment, field):
                        setattr(shipment, field, val)
                db.session.add(shipment)
                total_shipments += 1

            if total_shipments % 1000 == 0:
                db.session.flush()

        db.session.commit()

        return jsonify({
            'success': True,
            'created': created,
            'updated': updated,
            'total_shipments': total_shipments,
            'total_clients': len(created) + len(updated)
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'{type(e).__name__}: {str(e)}'}), 500
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ===================== EXPORT HELPERS =====================

def _fmt_dollar(val):
    """Format a value with $ sign if it looks numeric."""
    if not val:
        return val
    s = str(val).replace('$', '').replace(',', '').strip()
    try:
        num = float(s)
        if num == int(num):
            return f'${int(num):,}'
        return f'${num:,.2f}'
    except (ValueError, TypeError):
        return val

def _safe_float(val):
    """Parse a string to float, stripping $ and commas. Returns 0 on failure."""
    if not val:
        return 0.0
    s = str(val).replace('$', '').replace(',', '').strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _get_addcosts_display(rate):
    """Get addcosts names and values from extra_data, fallback to old columns."""
    extra = rate.get_extra()
    addcosts = extra.get('addcosts', [])
    if addcosts:
        names = '\n'.join(item.get('name', '') for item in addcosts if item.get('name'))
        values = '\n'.join(_fmt_dollar(item.get('value', '')) for item in addcosts if item.get('value'))
        return names, values
    # Fallback to old columns
    return rate.additional_costs or '', _fmt_dollar(rate.additional_costs_value) if rate.additional_costs_value else ''


# ===================== EXCEL GENERATORS =====================

def _generar_excel_cliente(cliente, hide_internal=False, tabla='all'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    wb = Workbook()
    ws = wb.active
    ws.title = cliente.nombre

    # Logo
    logo_path = get_logo_path()
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 180
        img.height = 50
        ws.add_image(img, 'A1')

    # Styles
    purple_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gray_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 4
    # Client name
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=cliente.nombre).font = Font(bold=True, size=14)
    row += 2

    custom_rates = cliente.get_custom_cols('rates')
    custom_airlines = cliente.get_custom_cols('airlines')
    custom_payments = cliente.get_custom_cols('payments')

    # Table 1: Airline Rates
    if tabla in ('all', 'rates'):
        if hide_internal:
            headers1 = ['Airline', 'Route', 'Transit Time', 'KG Availability', 'Date',
                        'Final Rate', 'Additional Costs', '', 'Notes']
            add_costs_col = 7
        else:
            headers1 = ['Airline', 'Route', 'Transit Time', 'KG Availability', 'Date',
                        'Net Rate', 'Operative', 'Net+OPS', 'Profit', 'Final Rate',
                        'Additional Costs', '', 'Notes']
            add_costs_col = 11
        headers1 += custom_rates
        for col, h in enumerate(headers1, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = purple_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=add_costs_col, end_row=row, end_column=add_costs_col + 1)
        row += 1
        sorted_rates = sorted(cliente.rates, key=lambda r: (r.airline or '').lower())
        data_start_row = row
        # Build group index per row (same airline group = same color)
        group_idx = 0
        row_group = []
        i = 0
        while i < len(sorted_rates):
            airline = (sorted_rates[i].airline or '').strip().lower()
            start = i
            while i < len(sorted_rates) and (sorted_rates[i].airline or '').strip().lower() == airline:
                row_group.append(group_idx)
                i += 1
            group_idx += 1
        for idx, r in enumerate(sorted_rates):
            addcost_names, addcost_values = _get_addcosts_display(r)
            if hide_internal:
                vals = [r.airline, r.route, r.transit_time, r.kg_availability, r.date,
                        _fmt_dollar(r.final_rate), addcost_names, addcost_values, r.notes]
            else:
                vals = [r.airline, r.route, r.transit_time, r.kg_availability, r.date,
                        _fmt_dollar(r.net_rate), _fmt_dollar(r.operative), _fmt_dollar(r.net_ops),
                        _fmt_dollar(r.profit), _fmt_dollar(r.final_rate),
                        addcost_names, addcost_values, r.notes]
            extra = r.get_extra()
            vals += [extra.get(c, '') for c in custom_rates]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = data_alignment
                if row_group[idx] % 2 == 1:
                    cell.fill = gray_fill
            row += 1
        # Merge airline cells for same-airline groups
        i = 0
        while i < len(sorted_rates):
            airline = (sorted_rates[i].airline or '').strip().lower()
            start = i
            while i < len(sorted_rates) and (sorted_rates[i].airline or '').strip().lower() == airline:
                i += 1
            if i - start > 1 and airline:
                ws.merge_cells(start_row=data_start_row + start, start_column=1,
                               end_row=data_start_row + i - 1, end_column=1)
        row += 1

    # Table 2: Current Status
    if tabla in ('all', 'status'):
        headers2 = ['Current Status', 'Next Flight', 'KG', 'Farms Deliver', 'Farm Maximum Delivery Time', 'Airline', 'All in Rate'] + custom_airlines
        for col, h in enumerate(headers2, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = purple_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        row += 1
        for idx, a in enumerate(cliente.airlines):
            extra = a.get_extra()
            vals = [a.current_status, a.proximo_vuelo, a.kg, a.entrega_fincas, a.hora_maxima, a.aerolinea, a.all_in_rate]
            vals += [extra.get(c, '') for c in custom_airlines]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = data_alignment
                if idx % 2 == 1:
                    cell.fill = gray_fill
            row += 1
        row += 1

    # Table 3: Payment (invoice-style layout per payment)
    if tabla in ('all', 'payment'):
        gold_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
        total_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')

        for p in cliente.payments:
            # --- Route header bar ---
            route_text = p.route or ''
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            cell = ws.cell(row=row, column=1, value=route_text)
            cell.fill = purple_fill
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = purple_fill
                ws.cell(row=row, column=c).border = thin_border
            ws.row_dimensions[row].height = 25
            row += 1

            # --- Headers row ---
            pay_headers = ['NET WEIGHT', 'VOLUME WEIGHT', 'RATE', 'WeightxRate',
                           'DUE AGENT', 'DUE CARRIER', 'CERTIFICATE', 'PHYTO', 'TOTAL']
            for col, h in enumerate(pay_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if h == 'TOTAL':
                    cell.fill = total_fill
                    cell.font = Font(bold=True, color='FFFFFF', size=9)
            row += 1

            # --- Data row ---
            nw = _safe_float(p.net_weight)
            vw = _safe_float(p.volume_weight)
            rate_val = _safe_float(p.rate)
            wxr = max(nw, vw) * rate_val
            da = _safe_float(p.pay_due_agent)
            dc = _safe_float(p.pay_due_carrier)
            cert = _safe_float(p.certificate)
            phyto = _safe_float(p.pay_phyto)
            total = wxr + da + dc + cert + phyto

            pay_vals = [
                p.net_weight or '',
                _fmt_dollar(p.volume_weight) if p.volume_weight else '',
                _fmt_dollar(p.rate) if p.rate else '',
                _fmt_dollar(str(round(wxr, 2))) if wxr else '',
                _fmt_dollar(p.pay_due_agent) if p.pay_due_agent else '',
                _fmt_dollar(p.pay_due_carrier) if p.pay_due_carrier else '',
                _fmt_dollar(p.certificate) if p.certificate else '',
                _fmt_dollar(p.pay_phyto) if p.pay_phyto else '',
                _fmt_dollar(str(round(total, 2))) if total else '',
            ]
            for col, v in enumerate(pay_vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.alignment = data_alignment
                cell.border = thin_border
                if col == 9:  # TOTAL column
                    cell.font = Font(bold=True, size=10)
            row += 1

            # --- NOTES section ---
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            cell = ws.cell(row=row, column=1, value='NOTES')
            cell.fill = purple_fill
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = purple_fill
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            notes_text = p.pay_notes or ''
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            cell = ws.cell(row=row, column=1, value=notes_text)
            cell.fill = gold_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(size=9)
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = gold_fill
                ws.cell(row=row, column=c).border = thin_border
            ws.row_dimensions[row].height = 45
            row += 2

        # --- ACCOUNT section (once, after all payments) ---
        if cliente.payments:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            cell = ws.cell(row=row, column=1, value='ACCOUNT')
            cell.fill = purple_fill
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = purple_fill
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            account_info = [
                ('BANK', 'Banco Pichincha'),
                ('SWIFT', 'PICHECEQXXX'),
                ('ACCOUNT', '2100339784'),
                ('COMPANY', 'FREIGHTWISE FORWARDING S.A'),
                ('TAXID', '1793231060012'),
                ('ADDRESS', 'Pedro de Alfaro y Francisco Ruiz'),
            ]
            for label, value in account_info:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                c1 = ws.cell(row=row, column=1, value=label)
                c1.font = Font(bold=True, size=9)
                c1.alignment = Alignment(horizontal='center', vertical='center')
                c1.fill = gray_fill
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = gray_fill
                    ws.cell(row=row, column=c).border = thin_border
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
                c2 = ws.cell(row=row, column=4, value=value)
                c2.font = Font(size=9)
                c2.alignment = Alignment(horizontal='center', vertical='center')
                for c in range(4, 10):
                    ws.cell(row=row, column=c).border = thin_border
                row += 1
            row += 1

    # Auto-width: increase cap to 50 and set minimum of 15
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = max(max_len + 4, 15)
        ws.column_dimensions[col[0].column_letter].width = min(width, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = f"_{tabla}" if tabla != 'all' else ''
    filename = f"current_status_{cliente.nombre}{suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _generar_excel_todos(clientes):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen Clientes'

    logo_path = get_logo_path()
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 180
        img.height = 50
        ws.add_image(img, 'A1')

    blue_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='Resumen de Clientes - Payment Status').font = Font(bold=True, size=14)
    row += 2

    headers = ['Cliente', 'Valor', 'Fecha', 'Credito', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    for c in clientes:
        last_payment = c.payments[-1] if c.payments else None
        vals = [
            c.nombre,
            last_payment.valor if last_payment else '',
            last_payment.fecha if last_payment else '',
            last_payment.credito if last_payment else '',
            c.estado.capitalize()
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if col == 5:
                cell.font = Font(bold=True, color='059669' if c.estado == 'finalizado' else 'D97706')
        row += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = max(max_len + 4, 15)
        ws.column_dimensions[col[0].column_letter].width = min(width, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"resumen_clientes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ===================== PDF GENERATORS =====================

def _generar_pdf_cliente(cliente, hide_internal=False, tabla='all'):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=10*mm, rightMargin=10*mm)
    elements = []
    styles = getSampleStyleSheet()

    # Cell paragraph styles (wraps text automatically)
    cell_style = ParagraphStyle('CellCenter', parent=styles['Normal'],
                                fontSize=7, leading=9, alignment=TA_CENTER)
    cell_header = ParagraphStyle('CellHeader', parent=styles['Normal'],
                                 fontSize=7, leading=9, alignment=TA_CENTER,
                                 textColor=colors.white)

    def P(text, style=cell_style):
        return Paragraph(str(text or ''), style)

    def PH(text):
        return Paragraph(str(text or ''), cell_header)

    # Logo
    logo_path = get_logo_path()
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=120, height=35)
        elements.append(logo)
        elements.append(Spacer(1, 10))

    # Client name
    elements.append(Paragraph(f'<b>{cliente.nombre}</b>', styles['Title']))
    elements.append(Spacer(1, 10))

    purple = colors.HexColor('#7C3AED')

    common_style = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    custom_rates = cliente.get_custom_cols('rates')
    custom_airlines = cliente.get_custom_cols('airlines')
    custom_payments = cliente.get_custom_cols('payments')

    # Table 1: Airline Rates
    if tabla in ('all', 'rates'):
        extra_w = [20*mm] * len(custom_rates)
        sorted_rates = sorted(cliente.rates, key=lambda r: (r.airline or '').lower())
        if hide_internal:
            data1 = [[PH('Airline'), PH('Route'), PH('Transit'), PH('KG'), PH('Date'),
                       PH('Final Rate'), PH('Additional Costs'), PH(''), PH('Notes')] + [PH(c) for c in custom_rates]]
            add_span = (6, 0, 7, 0)
            col_widths1 = [40*mm, 38*mm, 22*mm, 18*mm, 28*mm, 20*mm, 28*mm, 20*mm, 55*mm] + extra_w
            for r in sorted_rates:
                extra = r.get_extra()
                addcost_names, addcost_values = _get_addcosts_display(r)
                data1.append([P(r.airline), P(r.route), P(r.transit_time), P(r.kg_availability), P(r.date),
                              P(_fmt_dollar(r.final_rate)), P(addcost_names), P(addcost_values), P(r.notes)]
                             + [P(extra.get(c, '')) for c in custom_rates])
        else:
            data1 = [[PH('Airline'), PH('Route'), PH('Transit'), PH('KG'), PH('Date'),
                       PH('Net Rate'), PH('Operative'), PH('Net+OPS'), PH('Profit'), PH('Final Rate'),
                       PH('Add. Costs'), PH(''), PH('Notes')] + [PH(c) for c in custom_rates]]
            add_span = (10, 0, 11, 0)
            col_widths1 = [28*mm, 28*mm, 18*mm, 14*mm, 22*mm, 14*mm, 14*mm, 14*mm, 14*mm, 16*mm, 22*mm, 16*mm, 37*mm] + extra_w
            for r in sorted_rates:
                extra = r.get_extra()
                addcost_names, addcost_values = _get_addcosts_display(r)
                data1.append([P(r.airline), P(r.route), P(r.transit_time), P(r.kg_availability), P(r.date),
                              P(_fmt_dollar(r.net_rate)), P(_fmt_dollar(r.operative)), P(_fmt_dollar(r.net_ops)),
                              P(_fmt_dollar(r.profit)), P(_fmt_dollar(r.final_rate)),
                              P(addcost_names), P(addcost_values), P(r.notes)]
                             + [P(extra.get(c, '')) for c in custom_rates])
        # Build merge spans and group-based backgrounds for same-airline groups
        airline_spans = []
        group_bg = []
        group_idx = 0
        i = 0
        while i < len(sorted_rates):
            airline = (sorted_rates[i].airline or '').strip().lower()
            start = i
            while i < len(sorted_rates) and (sorted_rates[i].airline or '').strip().lower() == airline:
                i += 1
            if i - start > 1 and airline:
                airline_spans.append(('SPAN', (0, start + 1), (0, i)))  # +1 for header row
            bg_color = colors.HexColor('#E2E8F0') if group_idx % 2 == 1 else colors.white
            for row_i in range(start, i):
                group_bg.append(('BACKGROUND', (0, row_i + 1), (-1, row_i + 1), bg_color))
            group_idx += 1
        if len(data1) > 1:
            t1 = Table(data1, repeatRows=1, colWidths=col_widths1)
            t1.setStyle(TableStyle([
                ('SPAN', (add_span[0], add_span[1]), (add_span[2], add_span[3])),
                ('BACKGROUND', (0, 0), (-1, 0), purple),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ] + group_bg + common_style + airline_spans))
            elements.append(t1)
            elements.append(Spacer(1, 15))

    # Table 2: Current Status
    if tabla in ('all', 'status'):
        data2 = [[PH('Current Status'), PH('Next Flight'), PH('KG'), PH('Farms Deliver'), PH('Farm Max Delivery'), PH('Airline'), PH('All in Rate')]
                  + [PH(c) for c in custom_airlines]]
        col_widths2 = [40*mm, 40*mm, 20*mm, 40*mm, 40*mm, 30*mm, 25*mm] + [30*mm] * len(custom_airlines)
        for a in cliente.airlines:
            extra = a.get_extra()
            data2.append([P(a.current_status), P(a.proximo_vuelo), P(a.kg), P(a.entrega_fincas), P(a.hora_maxima), P(a.aerolinea), P(a.all_in_rate)]
                         + [P(extra.get(c, '')) for c in custom_airlines])
        if len(data2) > 1:
            t2 = Table(data2, repeatRows=1, colWidths=col_widths2)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), purple),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E2E8F0')]),
            ] + common_style))
            elements.append(t2)
            elements.append(Spacer(1, 15))

    # Table 3: Payment (invoice-style, matching image layout)
    if tabla in ('all', 'payment'):
        gold = colors.HexColor('#FFF8E1')
        total_w = 260 * mm  # full page width for merged cells

        cell_bold = ParagraphStyle('CellBold', parent=styles['Normal'],
                                   fontSize=8, leading=10, alignment=TA_CENTER,
                                   textColor=colors.black)

        for p_item in cliente.payments:
            # --- Route header bar ---
            route_bar = [[Paragraph(f'<b>{p_item.route or ""}</b>', cell_header)]]
            rt = Table(route_bar, colWidths=[total_w])
            rt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), purple),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(rt)

            # --- Charges header + data ---
            nw = _safe_float(p_item.net_weight)
            vw = _safe_float(p_item.volume_weight)
            rate_v = _safe_float(p_item.rate)
            wxr = max(nw, vw) * rate_v
            da = _safe_float(p_item.pay_due_agent)
            dc = _safe_float(p_item.pay_due_carrier)
            cert = _safe_float(p_item.certificate)
            phyto = _safe_float(p_item.pay_phyto)
            total_v = wxr + da + dc + cert + phyto

            pay_h = ['NET WEIGHT', 'VOLUME WEIGHT', 'RATE', 'WeightxRate',
                     'DUE AGENT', 'DUE CARRIER', 'CERTIFICATE', 'PHYTO', 'TOTAL']
            pay_hdr = [PH(h) for h in pay_h[:-1]] + [Paragraph('<b>TOTAL</b>', cell_header)]
            pay_data = [
                P(p_item.net_weight or ''),
                P(_fmt_dollar(p_item.volume_weight) if p_item.volume_weight else ''),
                P(_fmt_dollar(p_item.rate) if p_item.rate else ''),
                P(_fmt_dollar(str(round(wxr, 2))) if wxr else ''),
                P(_fmt_dollar(p_item.pay_due_agent) if p_item.pay_due_agent else ''),
                P(_fmt_dollar(p_item.pay_due_carrier) if p_item.pay_due_carrier else ''),
                P(_fmt_dollar(p_item.certificate) if p_item.certificate else ''),
                P(_fmt_dollar(p_item.pay_phyto) if p_item.pay_phyto else ''),
                Paragraph(f'<b>{_fmt_dollar(str(round(total_v, 2))) if total_v else ""}</b>', cell_bold),
            ]
            cw = [29*mm, 29*mm, 29*mm, 29*mm, 29*mm, 29*mm, 29*mm, 29*mm, 29*mm]
            charges_t = Table([pay_hdr, pay_data], colWidths=cw)
            charges_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-2, 0), colors.white),
                ('BACKGROUND', (-1, 0), (-1, 0), purple),  # TOTAL header purple
                ('TEXTCOLOR', (-1, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ] + common_style))
            elements.append(charges_t)

            # --- NOTES section ---
            notes_hdr = [[Paragraph('<b>NOTES</b>', cell_header)]]
            nt = Table(notes_hdr, colWidths=[total_w])
            nt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), purple),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(nt)

            notes_body = [[Paragraph(p_item.pay_notes or '', cell_style)]]
            nb = Table(notes_body, colWidths=[total_w])
            nb.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), gold),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(nb)
            elements.append(Spacer(1, 15))

        # --- ACCOUNT section (once) ---
        if cliente.payments:
            acct_hdr = [[Paragraph('<b>ACCOUNT</b>', cell_header)]]
            at = Table(acct_hdr, colWidths=[total_w])
            at.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), purple),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(at)

            account_info = [
                ('BANK', 'Banco Pichincha'),
                ('SWIFT', 'PICHECEQXXX'),
                ('ACCOUNT', '2100339784'),
                ('COMPANY', 'FREIGHTWISE FORWARDING S.A'),
                ('TAXID', '1793231060012'),
                ('ADDRESS', 'Pedro de Alfaro y Francisco Ruiz'),
            ]
            acct_data = []
            for label, value in account_info:
                acct_data.append([
                    Paragraph(f'<b>{label}</b>', cell_style),
                    Paragraph(value, cell_style)
                ])
            acct_t = Table(acct_data, colWidths=[80*mm, 180*mm])
            acct_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E2E8F0')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ] + common_style))
            elements.append(acct_t)

    doc.build(elements)
    output.seek(0)
    suffix = f"_{tabla}" if tabla != 'all' else ''
    filename = f"current_status_{cliente.nombre}{suffix}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')


def _generar_pdf_todos(clientes):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    logo_path = get_logo_path()
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=120, height=35)
        elements.append(logo)
        elements.append(Spacer(1, 10))

    elements.append(Paragraph('<b>Resumen de Clientes - Payment Status</b>', styles['Title']))
    elements.append(Spacer(1, 15))

    blue = colors.HexColor('#3B82F6')
    data = [['Cliente', 'Valor', 'Fecha', 'Credito', 'Estado']]
    for c in clientes:
        last = c.payments[-1] if c.payments else None
        data.append([
            c.nombre,
            last.valor if last else '',
            last.fecha if last else '',
            last.credito if last else '',
            c.estado.capitalize()
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t)

    doc.build(elements)
    output.seek(0)
    filename = f"resumen_clientes_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')
